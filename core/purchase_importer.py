"""
Invoice import helpers for the purchase workflow.

This module has no Tkinter UI code. It parses PDF/CSV/Excel invoices into a
small, validated structure and can then populate an existing PurchasePage using
the same item format that the normal manual entry flow already uses.
"""
import csv
import json
import os
import re
import sys
import tkinter as tk
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from core.purchase_service import get_or_create_medicine
from core.layout_config import is_strip_count_type, get_type_measure_unit
from core.purchase_invoice_engine import (
    PROMPT_VERSION,
    apply_record_pricing,
    classify_header_field,
    detect_document_format,
    enrich_purchase_invoice,
    is_skip_row,
    merge_multiline_item_records,
    resolve_line_pricing,
    save_import_alias,
)
from core.medicine_type_detector import (
    enrich_invoice_medicine_types,
    resolve_medicine_type,
)


PDF_POSITIONAL_FIELDS = (
    "hsn_code", "gst_pct", "mrp", "name", "pack", "manufacturer",
    "batch", "expiry", "qty", "free_qty", "rate", "amount",
)

FIELD_ALIASES = {
    "name": [
        "medicine name", "product name", "item name", "name of product",
        "product", "medicine", "item", "description", "particulars",
    ],
    "batch": ["batch", "batch no", "batch no.", "batch number", "batchno"],
    "expiry": [
        "expiry", "exp date", "expiry date", "expdt", "exp dt", "exp.",
        "exp", "expd", "exp. date",
    ],
    "qty": ["qty", "quantity", "qnty", "pcs", "units"],
    "free_qty": ["free", "free qty", "free quantity", "fqty"],
    "rate": ["rate", "purchase rate", "ptr", "pur rate", "p.rate", "price", "unit rate"],
    "gst_pct": ["gst", "gst%", "gst %", "tax", "tax%", "tax %"],
    "hsn_code": ["hsn", "hsn code", "hsncode", "hsn/sac", "sac"],
    "mrp": ["mrp", "m.r.p.", "m.r.p", "m r p"],
    "amount": ["amount", "net amount", "value", "total", "gross amount"],
    "pack": ["pack", "packing", "pack size", "size"],
    "qty_unit": ["qty unit", "qty. unit", "q unit", "quantity unit", "uom"],
    "pkg_unit": ["pkg unit", "pkg. unit", "pack unit", "package unit", "packing unit"],
    "manufacturer": ["mfg", "mfr", "manufacturer", "company", "make"],
    "medicine_type": ["type", "medicine type", "product type"],
    "discount_pct": ["disc%", "discount%", "discount percent", "disc percent", "disc pct"],
    "disc_rupees": [
        "disc.", "disc amt", "discount amt", "discount amount", "disc amount",
    ],
}

# Shown on Purchase page when the bill has no batch / expiry — update before saving.
IMPORT_PLACEHOLDER_BATCH = "WITHOUT BATCH"
IMPORT_PLACEHOLDER_EXPIRY = "WITHOUT EXP"

TYPE_UNITS = {
    "syrup": "ml",
    "injection": "ml",
    "liquid": "ml",
    "liniment": "ml",
    "ointment": "gm",
    "powder": "gm",
    "gel": "gm",
    "granules": "gm",
    "vaccine": "ml",
    "injection - vial": "Vial",
}


class InvoiceParseError(Exception):
    """Raised when an invoice file cannot be parsed into purchase rows."""


@dataclass
class ImportedPurchaseItem:
    name: str = ""
    medicine_type: str = ""
    batch: str = ""
    expiry: str = ""
    qty: float = 0.0
    free_qty: float = 0.0
    rate: float = 0.0
    discount_pct: float = 0.0
    disc_column_value: float = 0.0
    disc_column_type: str = "ABSENT"
    amount_validated: bool = True
    gst_pct: float = 0.0
    mrp: float = 0.0
    hsn_code: str = ""
    manufacturer: str = ""
    schedule: str = ""
    content_drug: str = ""
    pack: str = ""
    amount: float = 0.0
    source_row: int = 0
    raw: Dict[str, Any] = field(default_factory=dict)
    issues: List[str] = field(default_factory=list)

    def validate(self) -> List[str]:
        issues = []
        if not self.name.strip():
            issues.append("Missing medicine name")
        batch_val = self.batch.strip()
        if not batch_val:
            issues.append("Missing batch number")
        exp_val = self.expiry.strip()
        if not exp_val:
            issues.append("Missing or invalid expiry")
        elif exp_val.upper() != IMPORT_PLACEHOLDER_EXPIRY:
            if not normalize_expiry(exp_val):
                issues.append("Missing or invalid expiry")
        if self.qty <= 0 and self.free_qty <= 0:
            issues.append("Quantity must be greater than zero")
        if self.rate < 0:
            issues.append("Rate cannot be negative")
        if self.gst_pct < 0 or self.gst_pct > 100:
            issues.append("GST percent must be between 0 and 100")
        if self.mrp < 0:
            issues.append("MRP cannot be negative")
        self.issues = issues
        return issues

    @property
    def is_valid(self) -> bool:
        return not self.validate()

    def merge_key(self) -> Tuple[Any, ...]:
        return (
            _clean_key(self.name),
            _clean_key(self.batch),
            self.expiry,
            round(float(self.rate or 0), 4),
            round(float(self.gst_pct or 0), 4),
            round(float(self.mrp or 0), 4),
            _clean_key(self.hsn_code),
            _clean_key(self.manufacturer),
        )


@dataclass
class PurchaseInvoice:
    supplier_name: str = ""
    supplier_address: str = ""
    supplier_phone: str = ""
    supplier_gstin: str = ""
    supplier_dl: str = ""
    invoice_number: str = ""
    invoice_date: str = ""
    items: List[ImportedPurchaseItem] = field(default_factory=list)
    source_path: str = ""
    source_type: str = ""
    parser: str = ""
    invoice_total: float = 0.0
    document_format: str = ""
    product_discount: float = 0.0
    cash_discount: float = 0.0
    total_cgst: float = 0.0
    total_sgst: float = 0.0
    footer_gst_authoritative: bool = False
    taxable_amount: float = 0.0
    gross_amount: float = 0.0
    line_gross: float = 0.0
    discount_base: float = 0.0
    cash_discount_pct: float = 0.0
    product_discount_pct: float = 0.0
    item_discount_total: float = 0.0
    round_off: float = 0.0
    confidence_scores: Dict[str, Any] = field(default_factory=dict)
    validation: Dict[str, Any] = field(default_factory=dict)
    review_flags: List[str] = field(default_factory=list)
    issues: List[str] = field(default_factory=list)
    raw_text: str = ""

    @property
    def item_count(self) -> int:
        return len(self.items)

    @property
    def preview_amount(self) -> float:
        total = 0.0
        for item in self.items:
            taxable = float(item.amount or (item.qty * item.rate) or 0)
            total += taxable + (taxable * float(item.gst_pct or 0) / 100)
        return round(total, 2)


def parse_purchase_pdf(path: str) -> PurchaseInvoice:
    """Parse a PDF invoice using pdfplumber, camelot, then tabula-py fallback."""
    _ensure_file(path)
    raw_text_parts: List[str] = []
    table_rows: List[List[Any]] = []
    parsers_used: List[str] = []
    issues: List[str] = []

    try:
        rows, text = _parse_pdf_with_pdfplumber(path)
        if rows:
            table_rows.extend(rows)
            parsers_used.append("pdfplumber")
        if text:
            raw_text_parts.append(text)
    except ImportError:
        issues.append("pdfplumber is not installed")
    except Exception as exc:
        issues.append("pdfplumber failed: {}".format(exc))

    if not table_rows:
        try:
            rows = _parse_pdf_with_camelot(path)
            if rows:
                table_rows.extend(rows)
                parsers_used.append("camelot")
        except ImportError:
            issues.append("camelot is not installed")
        except Exception as exc:
            issues.append("camelot failed: {}".format(exc))

    if not table_rows:
        try:
            rows = _parse_pdf_with_tabula(path)
            if rows:
                table_rows.extend(rows)
                parsers_used.append("tabula-py")
        except ImportError:
            issues.append("tabula-py is not installed")
        except Exception as exc:
            issues.append("tabula-py failed: {}".format(exc))

    if not table_rows and not raw_text_parts:
        try:
            text = _parse_pdf_with_builtin_text(path)
            if text:
                raw_text_parts.append(text)
                parsers_used.append("built-in PDF text")
        except Exception as exc:
            issues.append("built-in PDF text failed: {}".format(exc))

    raw_text = "\n".join(part for part in raw_text_parts if part)

    marg_invoice = _try_parse_marg_erp_invoice(raw_text, path, issues, parsers_used)
    if marg_invoice is not None:
        if parsers_used:
            marg_invoice.parser = "{}, {}".format(
                marg_invoice.parser, ", ".join(parsers_used)
            )
        marg_invoice.source_path = path
        marg_invoice.source_type = "pdf"
        marg_invoice.raw_text = raw_text
        marg_invoice.document_format = detect_document_format(path, table_rows, raw_text)
        combined_issues = list(marg_invoice.issues) + list(issues)
        if parsers_used:
            combined_issues = [
                issue for issue in combined_issues if "is not installed" not in issue
            ]
        marg_invoice.issues = combined_issues
        _finalize_parsed_invoice(marg_invoice)
        _apply_ganesh_line_corrections(marg_invoice)
        return marg_invoice

    tuljai_invoice = _try_parse_tuljai_invoice(raw_text, table_rows, issues)
    if tuljai_invoice is not None:
        if parsers_used:
            tuljai_invoice.parser = "{}, {}".format(
                tuljai_invoice.parser, ", ".join(parsers_used)
            )
        tuljai_invoice.source_path = path
        tuljai_invoice.source_type = "pdf"
        tuljai_invoice.raw_text = raw_text
        tuljai_invoice.document_format = detect_document_format(path, table_rows, raw_text)
        combined_issues = list(tuljai_invoice.issues) + list(issues)
        if parsers_used:
            combined_issues = [issue for issue in combined_issues if "is not installed" not in issue]
        tuljai_invoice.issues = combined_issues
        _finalize_parsed_invoice(tuljai_invoice)
        _apply_ganesh_line_corrections(tuljai_invoice)
        return tuljai_invoice

    records = _records_from_table_rows(table_rows, issues)
    if not records and raw_text:
        records = _records_from_text(raw_text)
    records = _merge_broken_records(records)

    items = _items_from_records(records, issues)
    items = merge_duplicate_items(items, issues)
    details = extract_supplier_details(raw_text)
    if table_rows:
        details = _merge_supplier_details(
            details, _extract_supplier_from_split_header_table(table_rows, raw_text)
        )

    if not items:
        hint = "No purchase rows were found in this PDF."
        if issues:
            hint += " Parser messages: " + "; ".join(issues[:3])
        raise InvoiceParseError(hint)
    if parsers_used:
        issues = [issue for issue in issues if "is not installed" not in issue]

    invoice = _invoice_from_details(
        details,
        items,
        path,
        "pdf",
        ", ".join(parsers_used) or "text",
        _extract_total_amount(raw_text, items),
        issues,
        raw_text,
    )
    invoice.document_format = detect_document_format(path, table_rows, raw_text)
    _finalize_parsed_invoice(invoice)
    _apply_ganesh_line_corrections(invoice)
    return invoice


def parse_purchase_excel(path: str) -> PurchaseInvoice:
    """Parse CSV, XLSX, or XLS invoices with flexible column detection."""
    _ensure_file(path)
    ext = os.path.splitext(path)[1].lower()
    if ext not in (".csv", ".xlsx", ".xls"):
        raise InvoiceParseError("Unsupported import file type: {}".format(ext))

    rows = _read_tabular_file(path, ext)
    if not rows:
        raise InvoiceParseError("The invoice file is empty.")

    vendor_invoice = _parse_h_t_f_invoice(path, rows, ext)
    if vendor_invoice is not None:
        vendor_invoice.document_format = "EDI_CSV"
        _finalize_parsed_invoice(vendor_invoice)
        _apply_ganesh_line_corrections(vendor_invoice)
        return vendor_invoice

    text = _rows_to_text(rows[:40])
    issues: List[str] = []
    header_idx, mapping = _find_header_map(rows)
    records: List[Dict[str, Any]] = []

    if mapping:
        for row_no, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            if _row_is_blank(row) or _looks_like_total_row(row):
                continue
            rec = {"source_row": row_no}
            for field_name, idx in mapping.items():
                rec[field_name] = _cell_at(row, idx)
            if _record_has_content(rec) and not _looks_like_non_item_name(rec.get("name")):
                records.append(rec)
    else:
        issues.append("No clear header row found; tried positional invoice parsing.")
        records = _records_from_table_rows(rows, issues)

    records = _merge_broken_records(records)
    items = _items_from_records(records, issues)
    items = merge_duplicate_items(items, issues)
    details = extract_supplier_details(text)
    if _looks_like_tuljai_invoice(text, rows):
        details = _enhance_tuljai_header_details(text, details)

    if not items:
        raise InvoiceParseError(
            "No medicine rows were found. Check that the file has medicine, batch, "
            "expiry, quantity and rate columns."
        )

    invoice = _invoice_from_details(
        details,
        items,
        path,
        ext.lstrip("."),
        "csv" if ext == ".csv" else "excel",
        _extract_total_amount(text, items),
        issues,
        text,
    )
    invoice.document_format = detect_document_format(path, rows, text)
    _finalize_parsed_invoice(invoice)
    _apply_ganesh_line_corrections(invoice)
    return invoice


def _finalize_parsed_invoice(invoice: PurchaseInvoice, conn: Any = None) -> None:
    """Totals/validation plus medicine type detection on all lines."""
    enrich_purchase_invoice(invoice)
    enrich_invoice_medicine_types(invoice, conn=conn)


def _type_from_record(
    rec: Dict[str, Any],
    conn: Any = None,
    available_types: Optional[Iterable[str]] = None,
) -> str:
    name = _clean_cell(rec.get("name"))
    pack = _clean_cell(rec.get("pack"))
    qty_unit = _clean_cell(rec.get("qty_unit") or rec.get("unit"))
    pkg_unit = _clean_cell(rec.get("pkg_unit") or rec.get("pack"))
    explicit = _clean_cell(rec.get("medicine_type"))
    if explicit:
        return explicit
    return resolve_medicine_type(
        conn=conn,
        name=name,
        pack=pack,
        qty_unit=qty_unit,
        pkg_unit=pkg_unit,
        bill_text=" ".join(filter(None, [name, pack, qty_unit, pkg_unit])),
        available_types=available_types,
        save_learned=bool(conn),
    )


def is_missing_import_batch(batch: Any) -> bool:
    """True when supplier bill has no batch (**, blank, etc.)."""
    b = _clean_cell(batch).upper()
    if not b:
        return True
    if b in ("**", "*", "-", "NA", "N/A", "NONE", "NIL", "XX", "X"):
        return True
    if re.match(r"^\*+$", b):
        return True
    return False


def apply_import_placeholders(item: ImportedPurchaseItem) -> None:
    """Fill missing batch/expiry so rows can load on the Purchase page for manual edit."""
    if is_missing_import_batch(item.batch):
        item.batch = IMPORT_PLACEHOLDER_BATCH
    if not normalize_expiry(item.expiry):
        item.expiry = IMPORT_PLACEHOLDER_EXPIRY


def apply_import_placeholders_to_items(items: Sequence[ImportedPurchaseItem]) -> None:
    for item in items:
        apply_import_placeholders(item)


def normalize_expiry(value: Any) -> str:
    """Normalize expiry values to MM/YY. Returns an empty string if invalid."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return "{:02d}/{:02d}".format(value.month, value.year % 100)
    if isinstance(value, date):
        return "{:02d}/{:02d}".format(value.month, value.year % 100)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if value > 20000:
            try:
                dt = datetime(1899, 12, 30) + timedelta(days=int(value))
                return "{:02d}/{:02d}".format(dt.month, dt.year % 100)
            except Exception:
                return ""

    text = _clean_cell(value).upper()
    if not text:
        return ""
    text = re.sub(r"\b(EXPIRY|EXPDT|EXP|DATE|DT)\b", "", text).strip()
    text = text.replace(".", "/").replace("-", "/")
    text = re.sub(r"\s+", " ", text)

    digits_only = re.sub(r"\D", "", text)
    if len(digits_only) == 8:
        day, month, year = int(digits_only[:2]), int(digits_only[2:4]), int(digits_only[4:])
        if 1 <= month <= 12:
            return _format_mmyy(str(month), str(year))

    m = re.match(r"^(\d{1,2})/(\d{2,4})$", text)
    if m:
        return _format_mmyy(m.group(1), m.group(2))

    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", text)
    if m:
        first, second, year = int(m.group(1)), int(m.group(2)), m.group(3)
        month = second if first > 12 or second <= 12 else first
        return _format_mmyy(str(month), year)

    m = re.match(r"^(\d{4})/(\d{1,2})(?:/\d{1,2})?$", text)
    if m:
        return _format_mmyy(m.group(2), m.group(1))

    month_map = {
        "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
        "JUL": 7, "AUG": 8, "SEP": 9, "SEPT": 9, "OCT": 10,
        "NOV": 11, "DEC": 12,
    }
    m = re.search(
        r"\b(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|SEPT|OCT|NOV|DEC)[A-Z]*\s*/?\s*(\d{2,4})\b",
        text,
    )
    if m:
        return _format_mmyy(str(month_map[m.group(1)]), m.group(2))
    return ""


def extract_supplier_details(source_text: Any) -> Dict[str, str]:
    """Extract supplier header fields used by the purchase page."""
    text = _clean_text(str(source_text or ""))
    lines = [_clean_cell(line) for line in text.splitlines() if _clean_cell(line)]
    result = {
        "supplier_name": "",
        "supplier_address": "",
        "supplier_phone": "",
        "supplier_gstin": "",
        "supplier_dl": "",
        "invoice_number": "",
        "invoice_date": "",
    }

    for line in lines[:80]:
        supplier = _extract_supplier_from_line(line)
        if supplier and not result["supplier_name"]:
            result["supplier_name"] = supplier
        inv = _extract_invoice_no_from_line(line)
        if inv and not result["invoice_number"]:
            result["invoice_number"] = inv
        dt = _extract_date_from_line(line)
        if dt and not result["invoice_date"]:
            result["invoice_date"] = dt

    if not result["supplier_name"]:
        result["supplier_name"] = _extract_supplier_name(lines)

    # Generic fallback for unseen suppliers:
    # many invoices print "For : <Supplier Name>" near signature block.
    if not result["supplier_name"]:
        m = re.search(
            r"\bFor\s*[:\-]\s*([A-Z][A-Z0-9&\.\-\/\s]{3,80})",
            text,
            flags=re.IGNORECASE,
        )
        if m:
            result["supplier_name"] = _clean_cell(m.group(1))

    gstin = _extract_gstin(text)
    if gstin:
        result["supplier_gstin"] = gstin

    phone = _extract_phone(text)
    if phone:
        result["supplier_phone"] = phone

    dl_numbers = _extract_dl_numbers(text)
    if dl_numbers:
        result["supplier_dl"] = dl_numbers

    address = _extract_supplier_address(text, result["supplier_name"])
    if address and not _looks_like_bad_supplier_address(address):
        result["supplier_address"] = address

    result["supplier_name"] = _sanitize_supplier_name(result.get("supplier_name", ""))
    return result


def _sanitize_supplier_name(name: str) -> str:
    """Remove footer/jurisdiction/amount-in-words noise from supplier name."""
    value = _clean_cell(name)
    if not value:
        return ""

    # Keep only the right side when header accidentally contains "... For : NAME"
    m = re.search(r"\bFor\s*[:\-]\s*([A-Z][A-Z0-9&\.\-\/\s]{3,80})", value, flags=re.IGNORECASE)
    if m:
        value = _clean_cell(m.group(1))

    # Trim common legal/footer phrases that get merged into the name
    value = re.split(r"\bSubject to\b", value, maxsplit=1, flags=re.IGNORECASE)[0]
    value = re.split(r"\bJurisdiction\b", value, maxsplit=1, flags=re.IGNORECASE)[0]
    value = re.split(r"\bAmount in words\b", value, maxsplit=1, flags=re.IGNORECASE)[0]

    # Remove leading "For" if present
    value = re.sub(r"^\s*For\s*[:\-]?\s*", "", value, flags=re.IGNORECASE)

    # Normalize whitespace and punctuation tails
    value = _clean_cell(value).strip(" :-,")

    # MARG ERP: long header line merged into supplier name
    value = re.split(
        r"\s+SWAMI\s+SAMARTH\s+MEDICAL\s+AND\s+AGENCY",
        value,
        maxsplit=1,
        flags=re.IGNORECASE,
    )[0].strip()
    value = re.split(r"\s+SHOP\s+NO\.", value, maxsplit=1, flags=re.IGNORECASE)[0].strip()
    value = re.split(r"\s+CTS\s+NO\.", value, maxsplit=1, flags=re.IGNORECASE)[0].strip()
    if value.upper().startswith("M/S "):
        value = value[4:].strip()

    # Avoid returning accidental sentence-like text as a supplier name
    lower = value.lower()
    blocked = (
        "rupees only",
        "subject to",
        "jurisdiction",
        "amount in words",
        "bank details",
        "declaration",
        "ground floor",
        "plot no",
    )
    if any(token in lower for token in blocked):
        return ""
    if len(value) > 72:
        value = value[:72].rsplit(" ", 1)[0].strip()
    return value


def _invoice_from_details(
    details: Dict[str, str],
    items: List[ImportedPurchaseItem],
    source_path: str,
    source_type: str,
    parser: str,
    invoice_total: float,
    issues: List[str],
    raw_text: str,
) -> PurchaseInvoice:
    return PurchaseInvoice(
        supplier_name=details.get("supplier_name", ""),
        supplier_address=details.get("supplier_address", ""),
        supplier_phone=details.get("supplier_phone", ""),
        supplier_gstin=details.get("supplier_gstin", ""),
        supplier_dl=details.get("supplier_dl", ""),
        invoice_number=details.get("invoice_number", ""),
        invoice_date=details.get("invoice_date", ""),
        items=items,
        source_path=source_path,
        source_type=source_type,
        parser=parser,
        invoice_total=invoice_total,
        issues=issues,
        raw_text=raw_text,
    )


def import_into_purchase_page(
    purchase_page: Any,
    invoice: PurchaseInvoice,
    items: Optional[Sequence[ImportedPurchaseItem]] = None,
    replace_existing: bool = True,
) -> Dict[str, Any]:
    """Populate an existing PurchasePage with imported invoice data."""
    if purchase_page is None:
        raise ValueError("Purchase page is not available.")
    selected_items = list(items if items is not None else invoice.items)
    apply_import_placeholders_to_items(selected_items)
    invalid = [item for item in selected_items if not item.is_valid]
    if invalid:
        first = invalid[0]
        raise ValueError(
            "Cannot import invalid rows. Row {}: {}".format(
                first.source_row or "?", "; ".join(first.issues)
            )
        )

    supplier_name = (invoice.supplier_name or "").strip()
    bill_number = (invoice.invoice_number or "").strip()
    purchase_date = normalize_invoice_date(invoice.invoice_date)

    if supplier_name:
        purchase_page.supplier_name.set(supplier_name)
    _set_entry(purchase_page.bill_number, bill_number)
    _set_entry(purchase_page.purchase_date, purchase_date or datetime.now().strftime("%Y-%m-%d"))

    try:
        purchase_page.load_supplier_details()
    except Exception:
        pass

    if (invoice.supplier_address or "").strip():
        _set_entry(purchase_page.supplier_address, invoice.supplier_address)
    if (invoice.supplier_phone or "").strip():
        _set_entry(purchase_page.supplier_phone, invoice.supplier_phone)
    if (invoice.supplier_gstin or "").strip():
        _set_entry(purchase_page.supplier_gstin, invoice.supplier_gstin)
    if (invoice.supplier_dl or "").strip():
        _set_entry(purchase_page.supplier_dl, invoice.supplier_dl)

    available_types = getattr(purchase_page, "_med_types", [])
    converted: List[Dict[str, Any]] = []
    for item in selected_items:
        raw = item.raw or {}
        qty_u = str(raw.get("qty_unit") or raw.get("unit") or "")
        pkg_u = str(raw.get("pkg_unit") or item.pack or "")
        med_type = resolve_medicine_type(
            conn=purchase_page.conn,
            name=item.name.strip(),
            pack=item.pack or pkg_u,
            qty_unit=qty_u,
            pkg_unit=pkg_u,
            bill_text=" ".join(
                p for p in [item.name, item.pack, item.content_drug] if p
            ),
            available_types=available_types,
        )
        med_type = _match_type(med_type or item.medicine_type, available_types)
        medicine_id = get_or_create_medicine(
            purchase_page.conn,
            item.name.strip(),
            med_type,
            item.batch.strip(),
            item.expiry,
            float(item.gst_pct or 0),
            float(item.mrp or 0),
            float(item.rate or 0),
            item.manufacturer.strip(),
            item.hsn_code.strip(),
            item.schedule.strip(),
            item.content_drug.strip(),
        )
        from core.purchase_invoice_engine import format_discount_display
        disc_label = format_discount_display(
            item.discount_pct,
            item.disc_column_value,
            item.disc_column_type,
        )
        import_taxable = float(item.amount or (item.qty * item.rate) or 0)
        import_gst_amt = round(import_taxable * float(item.gst_pct or 0) / 100, 2)
        converted_item = {
            "medicine_id": medicine_id,
            "name": item.name.strip(),
            "type": med_type,
            "batch": item.batch.strip(),
            "expiry": item.expiry,
            "qty": float(item.qty or 0),
            "free_qty": float(item.free_qty or 0),
            "rate": float(item.rate or 0),
            "discount_pct": float(item.discount_pct or 0),
            "disc_column_value": float(item.disc_column_value or 0),
            "disc_column_type": str(item.disc_column_type or "ABSENT"),
            "discount_display": disc_label,
            "gst_pct": float(item.gst_pct or 0),
            "mrp": float(item.mrp or 0),
            "hsn_code": item.hsn_code.strip(),
            "pack": (item.pack or "").strip(),
            "manufacturer": item.manufacturer.strip(),
            "schedule": item.schedule.strip(),
            "content_drug": item.content_drug.strip(),
            # Pre-GST goods value = rate × billed qty (free qty is bonus only)
            "import_lock_values": True,
            "import_taxable": import_taxable,
            "import_gst_amt": import_gst_amt,
            "import_item_amount": round(import_taxable + import_gst_amt, 2),
        }
        _add_quantity_metadata(converted_item, item.pack)
        converted.append(converted_item)

    if replace_existing:
        purchase_page.purchase_items = converted
    else:
        purchase_page.purchase_items.extend(converted)
    purchase_page.purchase_items = _merge_purchase_page_items(purchase_page.purchase_items)

    # Mark page in import-as-is mode so totals use imported invoice values.
    setattr(purchase_page, "_import_bill_mode", True)
    line_gross = round(
        float(getattr(invoice, "line_gross", 0) or 0)
        or sum(float(it.amount or 0) for it in invoice.items),
        2,
    )
    gross_amount = float(getattr(invoice, "gross_amount", 0) or 0) or line_gross
    footer_cgst = float(getattr(invoice, "total_cgst", 0) or 0)
    footer_sgst = float(getattr(invoice, "total_sgst", 0) or 0)
    inv_net = float(invoice.invoice_total or 0)
    footer_gst = round(footer_cgst + footer_sgst, 2)
    setattr(
        purchase_page,
        "_import_invoice_summary",
        {
            "invoice_total": inv_net,
            "taxable_amount": float(getattr(invoice, "taxable_amount", 0) or 0),
            "gross_amount": gross_amount,
            "line_gross": line_gross,
            "discount_base": float(getattr(invoice, "discount_base", 0) or 0),
            "cash_discount_pct": float(getattr(invoice, "cash_discount_pct", 0) or 0),
            "product_discount_pct": float(getattr(invoice, "product_discount_pct", 0) or 0),
            "product_discount": float(getattr(invoice, "product_discount", 0) or 0),
            "cash_discount": float(getattr(invoice, "cash_discount", 0) or 0),
            "parsed_total_discount": round(
                float(getattr(invoice, "product_discount", 0) or 0)
                + float(getattr(invoice, "cash_discount", 0) or 0),
                2,
            ),
            "supplier_gross": float(getattr(invoice, "gross_amount", 0) or 0),
            "item_discount_total": float(getattr(invoice, "item_discount_total", 0) or 0),
            "edi_format": (
                "seema_legacy" if "seema_legacy" in (getattr(invoice, "parser", "") or "").lower()
                else "seema" if "seema" in (getattr(invoice, "parser", "") or "").lower()
                else "marg" if "marg" in (getattr(invoice, "parser", "") or "").lower()
                else ""
            ),
            "total_cgst": footer_cgst,
            "total_sgst": footer_sgst,
            "round_off": float(getattr(invoice, "round_off", 0) or 0),
            "footer_gst_authoritative": bool(
                getattr(invoice, "footer_gst_authoritative", False)
                or (footer_cgst > 0 and footer_sgst > 0 and inv_net > 0)
            ),
            "use_footer_totals": bool(
                inv_net > 0
                and footer_gst > 0
                and (gross_amount > 0 or line_gross > 0)
            ),
        },
    )
    total_disc = round(
        float(getattr(invoice, "product_discount", 0) or 0)
        + float(getattr(invoice, "cash_discount", 0) or 0),
        2,
    )
    if hasattr(purchase_page, "overall_discount"):
        _set_entry(purchase_page.overall_discount, "{:.2f}".format(total_disc))
    cash_pct = float(getattr(invoice, "cash_discount_pct", 0) or 0)
    if cash_pct and hasattr(purchase_page, "overall_discount_pct"):
        _set_entry(purchase_page.overall_discount_pct, "{:.2f}".format(cash_pct))
    elif hasattr(purchase_page, "sync_overall_discount_fields"):
        purchase_page.sync_overall_discount_fields('rupees')
    purchase_page.calculate_total()
    # calculate_total applies shared auto-rounding (same as manual purchase entry).
    purchase_page.update_items_tree()

    if supplier_name:
        save_import_alias("supplier_aliases", supplier_name, supplier_name)

    imported_count = len(converted)
    write_import_log(
        invoice,
        "imported",
        "{} item(s) imported into purchase page".format(imported_count),
    )
    return {
        "items_imported": imported_count,
        "items_on_page": len(purchase_page.purchase_items),
        "invoice_total": invoice.invoice_total,
        "preview_total": invoice.preview_amount,
    }


def normalize_invoice_date(value: Any) -> str:
    """Normalize invoice dates to YYYY-MM-DD, preferring Indian DD/MM/YYYY."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = _clean_cell(value)
    if not text:
        return ""
    text = text.replace(".", "/").replace("-", "/")

    m = re.search(r"(\d{4})/(\d{1,2})/(\d{1,2})", text)
    if m:
        return _valid_date(int(m.group(1)), int(m.group(2)), int(m.group(3)))

    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", text)
    if m:
        first, second, year_raw = int(m.group(1)), int(m.group(2)), m.group(3)
        year = int(year_raw) + (2000 if len(year_raw) == 2 else 0)
        day, month = first, second
        if first <= 12 and second > 12:
            month, day = first, second
        return _valid_date(year, month, day)

    for fmt in ("%d %b %Y", "%d %B %Y", "%b %d %Y", "%B %d %Y"):
        try:
            return datetime.strptime(text.title(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return ""


def _normalize_compact_date(value: Any) -> str:
    text = re.sub(r"\D", "", _clean_cell(value))
    if len(text) != 8:
        return ""
    day = int(text[:2])
    month = int(text[2:4])
    year = int(text[4:])
    return _valid_date(year, month, day)


def _normalize_compact_expiry(value: Any) -> str:
    text = re.sub(r"\D", "", _clean_cell(value))
    if len(text) == 8:
        return _format_mmyy(text[2:4], text[6:])
    if len(text) == 6:
        return _format_mmyy(text[2:4], text[4:])
    return ""


def merge_duplicate_items(
    items: Sequence[ImportedPurchaseItem],
    issues: Optional[List[str]] = None,
) -> List[ImportedPurchaseItem]:
    """Merge duplicate invoice rows by medicine/batch/expiry/rate/tax details."""
    merged: Dict[Tuple[Any, ...], ImportedPurchaseItem] = {}
    output: List[ImportedPurchaseItem] = []
    for item in items:
        if not item.name.strip() or not item.batch.strip() or not item.expiry.strip():
            output.append(item)
            continue
        key = item.merge_key()
        if key not in merged:
            merged[key] = item
            output.append(item)
            continue
        target = merged[key]
        target.qty = round(float(target.qty or 0) + float(item.qty or 0), 4)
        target.free_qty = round(float(target.free_qty or 0) + float(item.free_qty or 0), 4)
        target.amount = round(float(target.amount or 0) + float(item.amount or 0), 2)
        target.raw.setdefault("merged_source_rows", []).append(item.source_row)
        if issues is not None:
            issues.append(
                "Merged duplicate row {} into {}".format(
                    item.source_row or "?", target.name
                )
            )
    return output


def write_import_log(invoice: PurchaseInvoice, status: str, message: str = "") -> None:
    """Append a compact JSONL log entry for support/debugging."""
    try:
        log_dir = _config_dir()
        os.makedirs(log_dir, exist_ok=True)
        payload = {
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "status": status,
            "message": message,
            "source_path": invoice.source_path,
            "source_type": invoice.source_type,
            "parser": invoice.parser,
            "supplier": invoice.supplier_name,
            "invoice_number": invoice.invoice_number,
            "invoice_date": invoice.invoice_date,
            "item_count": len(invoice.items),
            "invoice_total": invoice.invoice_total,
            "preview_total": invoice.preview_amount,
            "issues": invoice.issues[:20],
        }
        with open(os.path.join(log_dir, "purchase_import.log"), "a", encoding="utf-8") as handle:
            handle.write(json.dumps(payload, ensure_ascii=True) + "\n")
    except Exception:
        pass


def _try_parse_marg_erp_invoice(
    raw_text: str,
    source_path: str,
    issues: List[str],
    parsers_used: Optional[List[str]] = None,
) -> Optional[PurchaseInvoice]:
    """Parse MARG ERP Nano chemist invoices (Sn./Mfg. or MRP/HSN line layouts)."""
    from core.marg_erp_parser import (
        enhance_marg_supplier_details,
        looks_like_marg_erp_invoice,
        pick_marg_item_records,
    )

    if not looks_like_marg_erp_invoice(raw_text):
        return None

    records, parser_label = pick_marg_item_records(raw_text)
    if not records:
        return None

    records = _merge_broken_records(records)
    local_issues: List[str] = []
    items = _items_from_records(records, local_issues)
    issues.extend(local_issues)
    items = merge_duplicate_items(items, issues)
    if not items:
        return None

    details = extract_supplier_details(raw_text)
    details = enhance_marg_supplier_details(raw_text, details)

    invoice = _invoice_from_details(
        details,
        items,
        source_path,
        "pdf",
        parser_label,
        _extract_total_amount(raw_text, items),
        issues,
        raw_text,
    )
    from core.marg_erp_parser import extract_marg_erp_footer_discounts
    marg_totals = extract_marg_erp_footer_discounts(raw_text)
    if marg_totals.get("product_discount"):
        invoice.product_discount = marg_totals["product_discount"]
    if marg_totals.get("cash_discount"):
        invoice.cash_discount = marg_totals["cash_discount"]
    if marg_totals.get("subtotal") and not invoice.taxable_amount:
        invoice.taxable_amount = marg_totals["subtotal"]
    return invoice


def _try_parse_tuljai_invoice(
    raw_text: str,
    table_rows: Sequence[Sequence[Any]],
    issues: List[str],
) -> Optional[PurchaseInvoice]:
    """Parse Micropro/Tuljai-style invoices with M.R.P. + Product columns."""
    if not _looks_like_tuljai_invoice(raw_text, table_rows):
        return None

    records: List[Dict[str, Any]] = []
    parser = "tuljai"

    line_records = _records_from_tuljai_text_lines(raw_text)
    stacked_records = _records_from_stacked_tuljai_table(table_rows)
    bill_item_count = _extract_bill_item_count(raw_text)

    # Text-line parsing is reliable across page breaks; stacked columns drift when
    # pack/name wraps (REFILL rows) on multi-page Tuljai bills.
    if bill_item_count and len(line_records) >= bill_item_count:
        records = line_records
        parser = "tuljai-text-lines"
    elif line_records and stacked_records:
        records = _merge_tuljai_record_lists(line_records, stacked_records)
        parser = "tuljai-merged"
    elif line_records:
        records = line_records
        parser = "tuljai-text-lines"
    elif stacked_records:
        records = stacked_records
        parser = "tuljai-stacked-table"
    else:
        records = []
        parser = "tuljai"

    if not records:
        jammed = _records_from_jammed_tuljai_text(raw_text)
        if jammed:
            records.extend(jammed)
            parser = "tuljai-jammed-text"

    if not records:
        header_idx, mapping = _find_header_map(table_rows)
        if mapping and header_idx >= 0:
            for row_no, row in enumerate(table_rows[header_idx + 1:], start=header_idx + 2):
                if _row_is_blank(row) or _looks_like_total_row(row):
                    continue
                rec = {"source_row": row_no}
                for field_name, idx in mapping.items():
                    rec[field_name] = _cell_at(row, idx)
                if _record_has_content(rec):
                    records.append(rec)
            parser = "tuljai-header-map"

    if not records:
        return None

    records = _merge_broken_records(records)
    items = _items_from_records(records, issues)
    items = merge_duplicate_items(items, issues)
    if not items:
        return None

    if bill_item_count and len(items) < bill_item_count:
        issues.append(
            "Bill lists {} item(s) but only {} row(s) were parsed — review before saving.".format(
                bill_item_count, len(items)
            )
        )

    details = extract_supplier_details(raw_text)
    details = _enhance_tuljai_header_details(raw_text, details, table_rows)
    return _invoice_from_details(
        details,
        items,
        "",
        "",
        parser,
        _extract_total_amount(raw_text, items),
        list(issues),
        raw_text,
    )


def _looks_like_tuljai_invoice(raw_text: str, table_rows: Sequence[Sequence[Any]]) -> bool:
    text = _clean_text(raw_text or "").lower()
    if "m.r.p" in text and "product" in text and ("batch" in text or "exp" in text):
        return True
    for row in table_rows[:20]:
        if _is_tuljai_header_row(row):
            return True
    return False


def _is_tuljai_header_row(row: Sequence[Any]) -> bool:
    labels = [_normalize_header(cell) for cell in row if _clean_cell(cell)]
    if not labels:
        return False
    has_mrp = any("mrp" in label for label in labels)
    has_product = any(label in ("product", "productname", "nameofproduct") for label in labels)
    has_qty = any(label in ("qty", "quantity") for label in labels)
    return has_mrp and has_product and has_qty


def _tuljai_column_map(row: Sequence[Any]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for idx, cell in enumerate(row):
        label = _normalize_header(cell)
        if not label:
            continue
        if "mrp" in label and "mrp" not in mapping:
            mapping["mrp"] = idx
        elif label in ("qty", "quantity") and "qty" not in mapping:
            mapping["qty"] = idx
        elif label == "free" and "free_qty" not in mapping:
            mapping["free_qty"] = idx
        elif label in ("product", "productname", "nameofproduct") and "name" not in mapping:
            mapping["name"] = idx
        elif label in ("pack", "packing", "pkg", "pkgsize", "package", "pkgunit") and "pack" not in mapping:
            mapping["pack"] = idx
        elif label in ("mfg", "mfr", "manufacturer") and "manufacturer" not in mapping:
            mapping["manufacturer"] = idx
        elif label in ("hsn", "hsncode", "hsnc") and "hsn_code" not in mapping:
            mapping["hsn_code"] = idx
        elif label in ("exp", "expdt", "expiry", "expdate") and "expiry" not in mapping:
            mapping["expiry"] = idx
        elif label in ("batch", "batchno", "batchnumber") and "batch" not in mapping:
            mapping["batch"] = idx
        elif label in ("disc", "discount", "disc.") and "disc_rupees" not in mapping:
            mapping["disc_rupees"] = idx
        elif label in ("discpct", "discpercent", "discountpct") and "discount_pct" not in mapping:
            mapping["discount_pct"] = idx
        elif label == "rate" and "rate" not in mapping:
            mapping["rate"] = idx
        elif label in ("gst", "gstpercent", "gstpct") and "gst_pct" not in mapping:
            mapping["gst_pct"] = idx
        elif label in ("amount", "netamount", "value") and "amount" not in mapping:
            mapping["amount"] = idx
        else:
            field = _field_for_header(cell)
            if field and field not in mapping:
                mapping[field] = idx
    return mapping


def _records_from_stacked_tuljai_table(table_rows: Sequence[Sequence[Any]]) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    for idx, row in enumerate(table_rows):
        if not _is_tuljai_header_row(row):
            continue
        if idx + 1 >= len(table_rows):
            continue
        mapping = _tuljai_column_map(row)
        if "name" not in mapping:
            continue
        data_row = table_rows[idx + 1]
        columns: Dict[str, List[str]] = {}
        max_len = 0
        for field_name, col_idx in mapping.items():
            cell = _cell_at(data_row, col_idx)
            parts = [_clean_cell(part) for part in str(cell or "").splitlines() if _clean_cell(part)]
            if not parts:
                parts = [_clean_cell(cell)]
            columns[field_name] = parts
            max_len = max(max_len, len(parts))
        if max_len <= 1:
            continue
        if "free_qty" in columns and len(columns["free_qty"]) == 1:
            free_value = columns["free_qty"][0] or "0"
            columns["free_qty"] = ["0"] * max_len
            columns["free_qty"][0] = free_value
        block_records: List[Dict[str, Any]] = []
        for row_no in range(max_len):
            rec = {"source_row": idx + 2 + row_no}
            for field_name, parts in columns.items():
                if row_no < len(parts):
                    rec[field_name] = parts[row_no]
                elif field_name == "free_qty":
                    rec[field_name] = "0"
                elif parts:
                    rec[field_name] = parts[-1]
            if not _record_has_content(rec) or _looks_like_non_item_name(rec.get("name")):
                continue
            if _is_tuljai_refill_continuation(rec):
                if block_records:
                    prev = block_records[-1]
                    extra = _clean_cell(rec.get("pack")) or _clean_cell(rec.get("name"))
                    if extra:
                        prev_pack = _clean_cell(prev.get("pack"))
                        prev["pack"] = "{} {}".format(prev_pack, extra).strip()
                continue
            block_records.append(rec)
        records.extend(block_records)
    return records


def _records_from_tuljai_text_lines(raw_text: str) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    started = False
    buffer = ""
    buffer_row = 0

    def flush_buffer():
        nonlocal buffer, buffer_row
        if not buffer:
            return
        rec = _parse_tuljai_item_line(buffer)
        if rec and not _looks_like_non_item_name(rec.get("name")):
            rec["source_row"] = buffer_row
            records.append(rec)
        buffer = ""

    for row_no, line in enumerate(raw_text.splitlines(), start=1):
        line = _clean_cell(line)
        if not line:
            continue
        low = line.lower()
        if "m.r.p" in low and "product" in low:
            started = True
            flush_buffer()
            continue
        if not started:
            continue
        if _looks_like_tuljai_footer_line(line):
            flush_buffer()
            continue
        if _is_tuljai_page_artifact(line):
            flush_buffer()
            continue
        if _looks_like_tuljai_item_line_start(line):
            flush_buffer()
            buffer = line
            buffer_row = row_no
            if _parse_tuljai_item_line(buffer):
                flush_buffer()
            continue
        if buffer:
            if _parse_tuljai_item_line(buffer):
                flush_buffer()
                if _is_tuljai_orphan_continuation(line) or _is_tuljai_page_artifact(line):
                    continue
            if _is_tuljai_orphan_continuation(line) or _is_tuljai_page_artifact(line):
                continue
            buffer = "{} {}".format(buffer, line)
    flush_buffer()
    return records


def _looks_like_tuljai_item_line_start(line: str) -> bool:
    return bool(re.match(r"^\d+\.\d{2}\s+\d", _clean_cell(line)))


def _is_tuljai_page_artifact(line: str) -> bool:
    """Page-break / carry-forward lines that must not merge into product rows."""
    low = _clean_cell(line).lower()
    if not low:
        return True
    if re.search(r"total\s+c/f|total\s+b/f|\bc/f\s*:|b/f\s*:", low):
        return True
    if "continued on" in low or "continued from" in low:
        return True
    if re.match(r"^page\s+\d", low):
        return True
    return False


def _is_tuljai_orphan_continuation(line: str) -> bool:
    """Wrapped pack/name fragment on the next line — ignore once the item row parsed."""
    low = _clean_cell(line).lower()
    if re.match(r"^refill\b", low):
        return True
    if low in ("tab", "tabs", "cap", "caps", "ml", "gm", "syp", "pow", "inj", "vial"):
        return True
    if re.match(r"^\d+\*\d", low):
        return True
    return False


def _is_tuljai_refill_continuation(rec: Dict[str, Any]) -> bool:
    name = _clean_cell(rec.get("name")).upper()
    if name == "REFILL" or name.startswith("REFILL "):
        return True
    if name in ("TAB", "TABS", "CAP", "CAPS") and not _clean_cell(rec.get("batch")):
        return True
    return False


def _merge_tuljai_record_lists(*sources: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    merged: Dict[Tuple[Any, ...], Dict[str, Any]] = {}
    order: List[Tuple[Any, ...]] = []
    for records in sources:
        for rec in records:
            if not _record_has_content(rec):
                continue
            if _looks_like_non_item_name(rec.get("name")):
                continue
            batch = _clean_key(rec.get("batch"))
            expiry = rec.get("expiry")
            rate = round(_to_float(rec.get("rate")), 4)
            name = _clean_key(rec.get("name"))
            if batch and expiry and rate > 0:
                key = (batch, expiry, rate)
            else:
                key = (name, batch, expiry, rate)
            if key not in merged:
                merged[key] = dict(rec)
                order.append(key)
                continue
            if _record_quality(rec) > _record_quality(merged[key]):
                merged[key] = dict(rec)
    return [merged[key] for key in order]


def _record_quality(rec: Dict[str, Any]) -> int:
    score = len(_clean_cell(rec.get("name")))
    if _clean_cell(rec.get("pack")):
        score += 8
    if _clean_cell(rec.get("hsn_code")):
        score += 4
    if _to_float(rec.get("qty")) > 0:
        score += 2
    return score


def _parse_tuljai_item_line(line: str) -> Optional[Dict[str, Any]]:
    line = _clean_cell(line)
    if not line or _looks_like_total_line(line):
        return None
    parts = line.split()
    if len(parts) < 10:
        return None
    if not re.match(r"^\d+\.\d{2}$", parts[-1]) and _to_float(parts[-1]) <= 0:
        return None

    amount = _to_float(parts[-1])
    gst_pct = _to_float(parts[-2])
    rate = _to_float(parts[-3])
    tail = parts[:-3]

    discount_pct = 0.0
    disc_raw = 0.0
    if tail and _looks_like_decimal(tail[-1]):
        test_tail = tail[:-1]
        if (
            len(test_tail) >= 3
            and _is_batch_token(test_tail[-1])
            and _is_expiry_token(test_tail[-2])
        ):
            disc_raw = _to_float(tail[-1])
            tail = test_tail

    if len(tail) < 6:
        return None
    batch = tail.pop()
    expiry = normalize_expiry(tail.pop())
    if not _is_batch_token(batch) or not expiry:
        return None

    hsn_parts: List[str] = []
    while tail and re.fullmatch(r"\d{3,8}", tail[-1]):
        hsn_parts.insert(0, tail.pop())
    hsn = "".join(hsn_parts)
    if len(tail) < 2:
        return None
    mfg = tail.pop()
    pack = tail.pop()
    head = tail

    if not _is_hsn_token(hsn):
        return None
    if len(head) < 2:
        return None

    mrp = _to_float(head[0])
    qty = _to_float(head[1])
    free_qty = 0.0
    product_start = 2
    if len(head) > 2 and _looks_like_decimal(head[2]):
        maybe_free = _to_float(head[2])
        if maybe_free <= max(qty, 1.0) and maybe_free < 100:
            free_qty = maybe_free
            product_start = 3
    name = " ".join(head[product_start:]).strip()
    if not name:
        return None

    pricing = resolve_line_pricing(mrp, rate, qty, amount or round(qty * rate, 2), disc_raw, "disc_rupees")

    return {
        "name": name,
        "pack": pack,
        "manufacturer": mfg,
        "hsn_code": hsn,
        "expiry": expiry,
        "batch": batch,
        "disc_rupees": pricing["disc_column_value"],
        "disc_column_type": pricing["disc_column_type"],
        "discount_pct": pricing["discount_pct"],
        "rate": pricing["rate"],
        "gst_pct": gst_pct,
        "amount": pricing["amount"],
        "mrp": mrp,
        "qty": qty,
        "free_qty": free_qty,
        "amount_validated": pricing["amount_validated"],
    }


def _records_from_jammed_tuljai_text(raw_text: str) -> List[Dict[str, Any]]:
    """Best-effort parser for PDFs where table columns are merged into one line."""
    if "|" in raw_text:
        return []

    records: List[Dict[str, Any]] = []
    pattern = re.compile(
        r"(\d+\.?\d*)\s+"
        r"([A-Z][A-Z0-9 \.\-/\'\*]+?)\s+"
        r"(\d{4,8})"
        r"([\d\.]+)"
        r"\s+([A-Z0-9\-]+)\s+"
        r"(\d{2}/\d{2})"
        r"(\d+\.?\d*)?"
    )
    for match in pattern.finditer(raw_text):
        qty = _to_float(match.group(1))
        name = _clean_cell(match.group(2))
        hsn = match.group(3)
        numeric_blob = match.group(4)
        batch = match.group(5)
        expiry = normalize_expiry(match.group(6))
        if not name or not expiry or qty <= 0:
            continue

        numbers = re.findall(r"\d+\.\d{2}|\d+", numeric_blob)
        floats = [_to_float(num) for num in numbers if _to_float(num) > 0]
        mrp = floats[0] if floats else 0.0
        gst_pct = 0.0
        rate = 0.0
        amount = 0.0
        for num in floats[1:]:
            if num in (5, 12, 18, 28, 0):
                gst_pct = num
            elif amount == 0 and num >= qty:
                amount = num
            elif rate == 0:
                rate = num
        if rate == 0 and amount and qty:
            rate = round(amount / qty, 4)
        if amount == 0 and rate:
            amount = round(qty * rate, 2)

        rec = {
            "name": name,
            "qty": qty,
            "free_qty": 0.0,
            "batch": batch,
            "expiry": expiry,
            "hsn_code": hsn,
            "mrp": mrp,
            "gst_pct": gst_pct,
            "rate": rate,
            "amount": amount,
        }
        if _record_has_content(rec):
            records.append(rec)
    return records


def _merge_supplier_details(
    base: Dict[str, str], extra: Dict[str, str]
) -> Dict[str, str]:
    """Prefer richer/extra supplier fields from table-header extraction."""
    result = dict(base)
    for key, value in (extra or {}).items():
        if not (value or "").strip():
            continue
        current = (result.get(key) or "").strip()
        incoming = value.strip()
        if not current:
            result[key] = incoming
            continue
        if key == "supplier_address" and (
            _looks_like_bad_supplier_address(current) or len(incoming) > len(current)
        ):
            result[key] = incoming
        elif key == "supplier_dl" and not current:
            result[key] = incoming
        elif key == "supplier_name" and len(incoming) > len(current):
            result[key] = _sanitize_supplier_name(incoming)
    return result


def _looks_like_dl_address_noise(value: str) -> bool:
    low = _clean_cell(value).lower()
    return bool(re.match(r"^dl\s*no", low))


def _looks_like_bad_supplier_address(value: str) -> bool:
    low = _clean_cell(value).lower()
    if not low:
        return True
    if _looks_like_dl_address_noise(value):
        return True
    if value.count("|") >= 2 or re.search(r"\border\s*no\b", low):
        return True
    blocked = (
        "jurisdiction",
        "subject to",
        "rupees only",
        "for :",
        "amount in words",
        "net amt",
        "cash disc",
    )
    return any(token in low for token in blocked)


def _extract_supplier_from_split_header_table(
    table_rows: Sequence[Sequence[Any]],
    raw_text: str = "",
) -> Dict[str, str]:
    """
    Two-column GST invoice headers (supplier left cell, customer right cell).
    Used by Jai Ganesh Pharmavet and similar stacked PDF layouts.
    """
    result: Dict[str, str] = {}
    supplier_cell = ""

    for row in table_rows[:5]:
        for cell in row:
            text = str(cell or "").strip()
            if not text or len(text) < 25:
                continue
            first_line = text.splitlines()[0].strip()
            if re.search(r"\bTo\s*:", first_line, flags=re.IGNORECASE):
                continue
            upper = text.upper()
            if ("GSTIN" in upper or "DL NO" in upper) and any(
                token in upper
                for token in ("PHARMA", "PHARMAVET", "MEDICAL", "DISTRIBUTOR", "AGENCY")
            ):
                supplier_cell = text
                break
        if supplier_cell:
            break

    if not supplier_cell:
        m = re.search(
            r"^([^\n]+?)\s+To\s*:\s*",
            _clean_text(raw_text or ""),
            flags=re.IGNORECASE | re.MULTILINE,
        )
        if m:
            result["supplier_name"] = _sanitize_supplier_name(m.group(1))
        return result

    lines = [_clean_cell(line) for line in supplier_cell.splitlines() if _clean_cell(line)]
    if not lines:
        return result

    result["supplier_name"] = _sanitize_supplier_name(lines[0])
    address_parts: List[str] = []

    for line in lines[1:]:
        low = line.lower()
        if re.search(r"\bdl\s*no", low, flags=re.IGNORECASE):
            dl_value = _extract_dl_from_line(line)
            if dl_value:
                existing = result.get("supplier_dl", "")
                result["supplier_dl"] = ", ".join(
                    p for p in [existing, dl_value] if p
                ).strip(", ")
            continue
        if "gstin" in low:
            gstin = _extract_gstin(line)
            if gstin:
                result["supplier_gstin"] = gstin
            continue
        if "contact" in low or "phone" in low or "mob" in low:
            phone = _extract_phone(line)
            if phone:
                result["supplier_phone"] = phone
            continue
        if "pan" in low and "gstin" not in low:
            continue
        cleaned = re.sub(
            r"\s*State\s*Code\s*:.*$", "", line, flags=re.IGNORECASE
        ).strip(" ,")
        if cleaned and not _looks_like_dl_address_noise(cleaned):
            address_parts.append(cleaned)

    if address_parts:
        result["supplier_address"] = ", ".join(address_parts)
    return result


def _extract_dl_from_line(line: str) -> str:
    """Pull licence numbers from 'DL NO ...' header lines."""
    text = _clean_cell(line)
    m = re.search(
        r"DL\s*NO\.?\s*(.+?)(?:\s*,?\s*Inv\.|\s+Inv\.|$)",
        text,
        flags=re.IGNORECASE,
    )
    if not m:
        m = re.search(r"DL\s*NO\.?\s*(.+)$", text, flags=re.IGNORECASE)
    if not m:
        return ""
    value = _clean_cell(m.group(1)).strip(" ,:-")
    value = re.sub(r"\s*,\s*Inv\..*$", "", value, flags=re.IGNORECASE).strip(" ,:-")
    return value if len(value) >= 4 else ""


def _enhance_tuljai_header_details(
    raw_text: str,
    details: Dict[str, str],
    table_rows: Optional[Sequence[Sequence[Any]]] = None,
) -> Dict[str, str]:
    result = dict(details)
    if table_rows:
        result = _merge_supplier_details(
            result, _extract_supplier_from_split_header_table(table_rows, raw_text)
        )
    text = _clean_text(raw_text or "")

    m = re.search(
        r"Credit Memo\s*:?\s*(INV\s*\d+)",
        text,
        flags=re.IGNORECASE,
    )
    if not m:
        m = re.search(
            r"Invoice\s+No\.?\s*:?\s*(INV\s*\d+)",
            text,
            flags=re.IGNORECASE,
        )
    if not m:
        m = re.search(r"\b(INV\s*\d+)\b", text, flags=re.IGNORECASE)
    if m:
        result["invoice_number"] = re.sub(r"\s+", " ", _clean_cell(m.group(1)))

    m = re.search(
        r"Invoice\s+Date\s*:?\s*(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})",
        text,
        flags=re.IGNORECASE,
    )
    if m:
        result["invoice_date"] = normalize_invoice_date(m.group(1))

    supplier_patterns = (
        r"(TULJAI MEDICAL AGENCY)",
        r"(SEEMA FRUITS[\s\w&]*PHARMA[\s\w&]*DISTRIBUTORS)",
        r"^([A-Z][A-Z0-9 &\.\-/]{4,60}(?:MEDICAL AGENCY|PHARMA[^|\n]{0,40}DISTRIBUTORS?))",
    )
    for pattern in supplier_patterns:
        m = re.search(pattern, text, flags=re.MULTILINE | re.IGNORECASE)
        if m:
            candidate = _clean_cell(m.group(1))
            if len(candidate) >= 4 and "billed" not in candidate.lower():
                result["supplier_name"] = candidate
                break

    if not result.get("supplier_gstin"):
        gstin = _extract_gstin(text, supplier_side=True)
        if gstin:
            result["supplier_gstin"] = gstin
    if not result.get("supplier_phone"):
        phone = _extract_phone(text, supplier_side=True)
        if phone:
            result["supplier_phone"] = phone
    if not result.get("supplier_dl"):
        dl_numbers = _extract_dl_numbers(text, supplier_side=True)
        if dl_numbers:
            result["supplier_dl"] = dl_numbers
    if not result.get("supplier_address") or _looks_like_bad_supplier_address(
        result.get("supplier_address", "")
    ):
        address = _extract_supplier_address(text, result.get("supplier_name", ""))
        if address and not _looks_like_bad_supplier_address(address):
            result["supplier_address"] = address
    if not result.get("supplier_dl"):
        dl_numbers = _extract_dl_numbers(text)
        if dl_numbers:
            result["supplier_dl"] = dl_numbers
    return result


def _looks_like_tuljai_footer_line(line: str) -> bool:
    low = _clean_cell(line).lower()
    if not low:
        return False
    if low.startswith("message"):
        return True
    if "cgst %" in low and "sgst %" in low:
        return True
    if "amount in words" in low or "net amt" in low:
        return True
    if "grand total" in low or "gross amount" in low:
        return True
    if "declaration" in low and "invoice" in low:
        return True
    return False


def _looks_like_non_item_name(value: Any) -> bool:
    name = _clean_cell(value).lower()
    if not name:
        return True
    if name in ("refill", "tab", "ml", "gm", "pow", "syp", "cap", "inj", "message"):
        return True
    blocked = (
        "message", "cgst", "sgst", "taxable", "gross", "declaration",
        "amount in words", "net amt", "bank name", "authorised",
    )
    return any(token in name for token in blocked)


def _is_expiry_token(value: Any) -> bool:
    text = _clean_cell(value)
    return bool(re.fullmatch(r"\d{2}/\d{2}", text))


def _is_hsn_token(value: Any) -> bool:
    text = re.sub(r"\D", "", _clean_cell(value))
    return 4 <= len(text) <= 8


def _is_batch_token(value: Any) -> bool:
    text = _clean_cell(value)
    if not text:
        return False
    if _is_expiry_token(text):
        return False
    if re.fullmatch(r"\d+(?:\.\d+)?", text) and len(text) >= 3:
        return True
    return bool(re.fullmatch(r"[A-Z0-9][A-Z0-9\-/\.]{2,20}", text, flags=re.IGNORECASE))


def _looks_like_decimal(value: Any) -> bool:
    text = _clean_cell(value)
    return bool(re.fullmatch(r"\d+(?:\.\d+)?", text))


def _score_invoice_sheet_rows(rows: Sequence[Sequence[Any]]) -> int:
    if not rows:
        return 0
    header_idx, mapping = _find_header_map(rows)
    score = 0
    if mapping and "name" in mapping:
        score = len(mapping) * 2
        for row in rows[header_idx + 1:header_idx + 25]:
            rec = {field: _cell_at(row, idx) for field, idx in mapping.items()}
            if _record_has_content(rec):
                score += 1
    if any(_is_tuljai_header_row(row) for row in rows[:25]):
        score = max(score, 18)
    t_rows = sum(1 for row in rows if _clean_cell(_cell_at(row, 0)).upper() == "T")
    if t_rows:
        score = max(score, 20 + t_rows)
    return score


def _parse_pdf_with_pdfplumber(path: str) -> Tuple[List[List[Any]], str]:
    import pdfplumber

    rows: List[List[Any]] = []
    text_parts: List[str] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if page_text:
                text_parts.append(page_text)
            for table in page.extract_tables() or []:
                rows.extend(table or [])
    return rows, "\n".join(text_parts)


def _parse_pdf_with_camelot(path: str) -> List[List[Any]]:
    import camelot

    rows: List[List[Any]] = []
    for flavor in ("lattice", "stream"):
        try:
            tables = camelot.read_pdf(path, pages="all", flavor=flavor)
        except Exception:
            if flavor == "stream":
                raise
            continue
        for table in tables:
            rows.extend(_rows_from_dataframe(table.df))
        if rows:
            break
    return rows


def _parse_pdf_with_tabula(path: str) -> List[List[Any]]:
    import tabula

    dfs = tabula.read_pdf(path, pages="all", multiple_tables=True)
    rows: List[List[Any]] = []
    for df in dfs or []:
        rows.extend(_rows_from_dataframe(df))
    return rows


def _parse_pdf_with_builtin_text(path: str) -> str:
    """Extract simple text-only PDF streams without optional PDF packages."""
    import zlib

    data = open(path, "rb").read()
    chunks: List[str] = []
    for match in re.finditer(rb"stream\r?\n(.*?)\r?\nendstream", data, flags=re.S):
        stream_data = match.group(1)
        prefix = data[max(0, match.start() - 300):match.start()]
        if b"FlateDecode" in prefix:
            try:
                stream_data = zlib.decompress(stream_data)
            except Exception:
                continue
        stream_text = stream_data.decode("latin-1", errors="ignore")
        chunks.extend(_extract_pdf_text_strings(stream_text))
    return "\n".join(line for line in chunks if line.strip())


def _extract_pdf_text_strings(stream_text: str) -> List[str]:
    lines: List[str] = []
    current = ""
    idx = 0
    while idx < len(stream_text):
        if stream_text.startswith("T*", idx):
            if current.strip():
                lines.append(current.strip())
            current = ""
            idx += 2
            continue
        if stream_text[idx] == "(":
            value, next_idx = _read_pdf_literal_string(stream_text, idx)
            current += value
            idx = next_idx
            continue
        idx += 1
    if current.strip():
        lines.append(current.strip())
    return lines


def _read_pdf_literal_string(text: str, start: int) -> Tuple[str, int]:
    chars: List[str] = []
    idx = start + 1
    depth = 1
    while idx < len(text) and depth > 0:
        ch = text[idx]
        if ch == "\\":
            if idx + 1 >= len(text):
                break
            nxt = text[idx + 1]
            escape_map = {
                "n": "\n", "r": "\r", "t": "\t",
                "b": "\b", "f": "\f", "\\": "\\",
                "(": "(", ")": ")",
            }
            if nxt in escape_map:
                chars.append(escape_map[nxt])
                idx += 2
                continue
            if nxt in "\r\n":
                idx += 2
                if nxt == "\r" and idx < len(text) and text[idx] == "\n":
                    idx += 1
                continue
            octal = re.match(r"[0-7]{1,3}", text[idx + 1:idx + 4])
            if octal:
                chars.append(chr(int(octal.group(0), 8)))
                idx += 1 + len(octal.group(0))
                continue
            chars.append(nxt)
            idx += 2
            continue
        if ch == "(":
            depth += 1
            chars.append(ch)
        elif ch == ")":
            depth -= 1
            if depth:
                chars.append(ch)
        else:
            chars.append(ch)
        idx += 1
    return "".join(chars), idx


def _read_tabular_file(path: str, ext: str) -> List[List[Any]]:
    if ext == ".csv":
        return _read_csv_rows(path)
    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise InvoiceParseError("openpyxl is required for .xlsx import.") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            best_rows: List[List[Any]] = []
            best_score = 0
            for sheet in workbook.worksheets:
                rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
                score = _score_invoice_sheet_rows(rows)
                if score > best_score:
                    best_score = score
                    best_rows = rows
            if best_rows:
                return best_rows
            sheet = workbook.active
            return [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()

    try:
        import pandas as pd
    except ImportError as exc:
        raise InvoiceParseError("pandas and xlrd are required for .xls import.") from exc

    best_rows: List[List[Any]] = []
    best_score = 0
    workbook = pd.ExcelFile(path)
    try:
        for sheet_name in workbook.sheet_names:
            df = pd.read_excel(path, sheet_name=sheet_name, header=None)
            rows = _rows_from_dataframe(df)
            score = _score_invoice_sheet_rows(rows)
            if score > best_score:
                best_score = score
                best_rows = rows
    finally:
        workbook.close()
    if best_rows:
        return best_rows
    df = pd.read_excel(path, header=None)
    return _rows_from_dataframe(df)


def _read_csv_rows(path: str) -> List[List[Any]]:
    last_error = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            with open(path, "r", encoding=encoding, newline="") as handle:
                sample = handle.read(4096)
                handle.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample)
                except csv.Error:
                    dialect = csv.excel
                try:
                    return [row for row in csv.reader(handle, dialect)]
                except (csv.Error, ValueError):
                    handle.seek(0)
                    return [row for row in csv.reader(handle)]
        except UnicodeDecodeError as exc:
            last_error = exc
    raise InvoiceParseError("Could not read CSV encoding: {}".format(last_error))


# Seema / Micropro EDI CSV (~33 T columns, T[1]=='0'):
#   H[2]=invoice no, H[3]=invoice date (DDMMYYYY), H[16]=customer name
#   T[4]=product code, T[5]=name, T[6]=pack, T[7]=mfg, T[8]=batch, T[9]=expiry
#   T[10]=qty, T[11]=free qty, T[13]=PTR rate, T[14]=line gross, T[16]=MRP
#   T[20]=line CD%, T[21]=line PD%, T[26]=CGST%, T[27]=SGST%, T[28]=GST%
#   T[30]=HSN, T[31]=line amount (usually = T[14])
#   F[1]=supplier gross (internal), F[2]=total GST (≈3% of F[1] on Seema bills)
#   F[22]=round off, F[23]=net payable, F[24]=cash disc ₹, F[25]=product disc ₹
#   F[24] and F[25] are each the same % (usually 3%) of one shared discount base.
#
# Seema legacy EDI (~31 cols, empty T[1], product code in T[3]):
#   T[5]=name, T[6]=pack, T[7]=mfg, T[8]=batch, T[9]=expiry, T[10]=rate, T[12]=MRP
#   T[15]=qty, T[21]=line amount, T[22]=CD%, T[24]=GST%, T[26]=CGST%, T[30]=HSN
#   F[1]=taxable, F[4]=CGST ₹, F[6]=SGST ₹, F[20]=round, F[21]=net payable
#
# MARG ERP EDI CSV (~51 T columns, <MARGERP FORMAT> in H row):
#   H[2]=invoice no, H[3]=date, H[30]=customer, H[52]=format marker
#   T[5]=name, T[6]=pack, T[8]=batch, T[9]=expiry, T[12]=qty, T[14]=rate
#   T[20]=qty billed, T[23]=line discount ₹, T[25]=taxable amount, T[26]=GST ₹
#   T[38]=HSN, T[41]=line total (with GST)
#   F[1]=net payable (rounded), F[2]=total item discount ₹ (already in T[25])
#   F[9]=total GST


def _detect_edi_htf_format(row: Sequence[Any], header_text: str = "") -> str:
    """
    Distinguish EDI CSV variants:
      seema — Seema / Micropro (~33 cols, T[1]=='0')
      marg  — MARG ERP Nano export (~51 cols, alphanumeric T[1])
    """
    ncol = len(row or [])
    marker = _clean_cell(_cell_at(row, 1))
    if "MARGERP" in (header_text or "").upper():
        return "marg"
    if ncol >= 40 and marker and not re.match(r"^[\d.]+$", marker):
        return "marg"
    if ncol >= 28 and marker in ("0", "0.0"):
        return "seema"
    if ncol >= 28 and not marker and _clean_cell(_cell_at(row, 5)):
        return "seema_legacy"
    return "legacy"


def _edi_htf_line_amounts(
    row: Sequence[Any],
    fmt: str = "",
) -> Tuple[float, float, float, float, float, float, str]:
    """
    Map H/T/F T-row columns by vendor format.
    Returns rate, mrp, qty, free_qty, amount (pre-GST taxable), gst_pct, hsn_code.
    """
    fmt = fmt or _detect_edi_htf_format(row)
    ncol = len(row or [])

    if fmt == "marg":
        rate = _to_float(_cell_at(row, 14))
        mrp = _to_float(_cell_at(row, 16))
        qty = _to_float(_cell_at(row, 20))
        free_qty = _to_float(_cell_at(row, 21))
        amount = _to_float(_cell_at(row, 25))
        if not amount and rate and qty:
            disc = _to_float(_cell_at(row, 23))
            amount = round(max(0.0, rate * qty - disc), 2)
        gst_amt = _to_float(_cell_at(row, 26))
        gst_pct = round(gst_amt / amount * 100, 2) if amount > 0 and gst_amt else 0.0
        hsn_code = _clean_cell(_cell_at(row, 38))
        return rate, mrp, qty, free_qty, amount, gst_pct, hsn_code

    if fmt == "seema":
        rate = _to_float(_cell_at(row, 13))
        mrp = _to_float(_cell_at(row, 16))
        qty = _to_float(_cell_at(row, 10))
        free_qty = _to_float(_cell_at(row, 11))
        amount = _to_float(_cell_at(row, 31)) or _to_float(_cell_at(row, 14))
        if not amount and rate and qty:
            amount = round(rate * qty, 2)
        gst_pct = _to_float(_cell_at(row, 28))
        if not gst_pct:
            # T[26]/T[27] are CGST%/SGST% slabs, not rupee amounts.
            gst_pct = _to_float(_cell_at(row, 26)) + _to_float(_cell_at(row, 27))
        hsn_code = _clean_cell(_cell_at(row, 30))
        return rate, mrp, qty, free_qty, amount, gst_pct, hsn_code

    if fmt == "seema_legacy":
        rate = _to_float(_cell_at(row, 10))
        mrp = _to_float(_cell_at(row, 12))
        qty = _to_float(_cell_at(row, 15))
        free_qty = 0.0
        amount = _to_float(_cell_at(row, 21)) or round(rate * qty, 2)
        if not amount and rate and qty:
            amount = round(rate * qty, 2)
        gst_pct = _to_float(_cell_at(row, 24))
        if not gst_pct:
            gst_pct = _to_float(_cell_at(row, 22)) + _to_float(_cell_at(row, 26))
        hsn_code = _clean_cell(_cell_at(row, 30))
        return rate, mrp, qty, free_qty, amount, gst_pct, hsn_code

    if ncol >= 25:
        rate = _to_float(_cell_at(row, 13)) or _to_float(_cell_at(row, 14))
        mrp = _to_float(_cell_at(row, 16))
        qty = _to_float(_cell_at(row, 10)) or _to_float(_cell_at(row, 20))
        free_qty = _to_float(_cell_at(row, 11)) or _to_float(_cell_at(row, 18))
        amount = _to_float(_cell_at(row, 31)) or _to_float(_cell_at(row, 14))
        if not amount and rate and qty:
            amount = round(rate * qty, 2)
        gst_pct = _to_float(_cell_at(row, 28)) or _to_float(_cell_at(row, 24))
        hsn_code = _clean_cell(_cell_at(row, 30))
        return rate, mrp, qty, free_qty, amount, gst_pct, hsn_code

    rate = _to_float(_cell_at(row, 10)) or _to_float(_cell_at(row, 11))
    mrp = _to_float(_cell_at(row, 12))
    qty = _to_float(_cell_at(row, 15))
    free_qty = _to_float(_cell_at(row, 16))
    amount = _to_float(_cell_at(row, 21))
    gst_pct = _to_float(_cell_at(row, 24))
    if not gst_pct:
        gst_pct = _to_float(_cell_at(row, 22)) + _to_float(_cell_at(row, 26))
    hsn_code = _clean_cell(_cell_at(row, 30)) if ncol > 30 else ""
    return rate, mrp, qty, free_qty, amount, gst_pct, hsn_code


def _derive_seema_discount_rates(
    cash_disc: float,
    prod_disc: float,
) -> Tuple[float, float, float]:
    """
    Seema F[24]/F[25] are cash/product discount rupees at a fixed % of one shared base.
    Returns (discount_base, cash_discount_pct, product_discount_pct).
    """
    cash_disc = round(float(cash_disc or 0), 2)
    prod_disc = round(float(prod_disc or 0), 2)
    if cash_disc <= 0 and prod_disc <= 0:
        return 0.0, 0.0, 0.0
    ref = cash_disc if cash_disc > 0 else prod_disc
    # Seema bills usually apply 3% CD + 3% PD on one shared base; prefer 3% first.
    for pct in (3.0, 2.5, 2.0, 1.5, 1.0, 3.5, 4.0, 5.0):
        base = round(ref * 100.0 / pct, 2)
        if base <= 0:
            continue
        cash_calc = round(base * pct / 100.0, 2)
        if abs(cash_calc - cash_disc) > 0.02:
            continue
        if prod_disc:
            prod_calc = round(base * pct / 100.0, 2)
            if abs(prod_disc - cash_disc) <= 0.02 and abs(prod_calc - prod_disc) <= 0.02:
                return base, pct, pct
            prod_pct = round(prod_disc * 100.0 / base, 2)
            return base, pct, prod_pct
        return base, pct, 0.0
    base = round(ref / 0.03, 2) if ref else 0.0
    cash_pct = round(cash_disc * 100.0 / base, 2) if base else 0.0
    prod_pct = round(prod_disc * 100.0 / base, 2) if base and prod_disc else 0.0
    return base, cash_pct, prod_pct


def _apply_edi_htf_footer(invoice: PurchaseInvoice, f_row: Sequence[Any], fmt: str) -> None:
    """Read bill-level totals from the F (footer) row."""
    if not f_row:
        return
    if fmt == "seema":
        line_gross = round(sum(float(it.amount or 0) for it in invoice.items), 2)
        footer_gross = _to_float(_cell_at(f_row, 1))
        total_gst = _to_float(_cell_at(f_row, 2))
        cash_disc = _to_float(_cell_at(f_row, 24))
        prod_disc = _to_float(_cell_at(f_row, 25))
        round_off = _to_float(_cell_at(f_row, 22))
        net = _to_float(_cell_at(f_row, 23))
        disc_base, cash_pct, prod_pct = _derive_seema_discount_rates(cash_disc, prod_disc)
        invoice.line_gross = line_gross
        invoice.discount_base = disc_base
        invoice.cash_discount_pct = cash_pct
        invoice.product_discount_pct = prod_pct
        if footer_gross:
            invoice.gross_amount = footer_gross
        elif line_gross:
            invoice.gross_amount = line_gross
        if cash_disc:
            invoice.cash_discount = cash_disc
        if prod_disc:
            invoice.product_discount = prod_disc
        if round_off:
            invoice.round_off = round_off
        if total_gst:
            invoice.total_cgst = round(total_gst / 2, 2)
            invoice.total_sgst = round(total_gst - invoice.total_cgst, 2)
            invoice.footer_gst_authoritative = True
        if net:
            invoice.invoice_total = net
        if net and total_gst:
            invoice.taxable_amount = round(net - total_gst - round_off, 2)
        return

    if fmt == "seema_legacy":
        line_gross = round(sum(float(it.amount or 0) for it in invoice.items), 2)
        footer_taxable = _to_float(_cell_at(f_row, 1))
        total_cgst = _to_float(_cell_at(f_row, 4))
        total_sgst = _to_float(_cell_at(f_row, 6))
        round_off = _to_float(_cell_at(f_row, 20))
        net = _to_float(_cell_at(f_row, 21))
        invoice.line_gross = line_gross
        invoice.gross_amount = footer_taxable or line_gross
        invoice.taxable_amount = footer_taxable or line_gross
        if round_off:
            invoice.round_off = round_off
        if total_cgst:
            invoice.total_cgst = total_cgst
        if total_sgst:
            invoice.total_sgst = total_sgst
        if total_cgst or total_sgst:
            invoice.footer_gst_authoritative = True
        if net:
            invoice.invoice_total = net
        return

    if fmt == "marg":
        line_taxable = round(sum(float(it.amount or 0) for it in invoice.items), 2)
        line_gst = round(
            sum(
                round(float(it.amount or 0) * float(it.gst_pct or 0) / 100, 2)
                for it in invoice.items
            ),
            2,
        )
        footer_net = _to_float(_cell_at(f_row, 1))
        footer_item_disc = _to_float(_cell_at(f_row, 2))
        footer_gst = _to_float(_cell_at(f_row, 9))
        total_gst = footer_gst or line_gst
        invoice.line_gross = line_taxable
        invoice.gross_amount = line_taxable
        invoice.taxable_amount = line_taxable
        if footer_item_disc:
            invoice.item_discount_total = footer_item_disc
        if total_gst:
            invoice.total_cgst = round(total_gst / 2, 2)
            invoice.total_sgst = round(total_gst - invoice.total_cgst, 2)
            invoice.footer_gst_authoritative = True
        pre_round = round(line_taxable + total_gst, 2)
        if footer_net:
            invoice.invoice_total = footer_net
            invoice.round_off = round(footer_net - pre_round, 2)
        else:
            from core.calc_engine import auto_round

            invoice.round_off = auto_round(pre_round)
            invoice.invoice_total = round(pre_round + invoice.round_off, 2)


def _parse_h_t_f_invoice(path: str, rows: Sequence[Sequence[Any]], ext: str) -> Optional[PurchaseInvoice]:
    """Parse EDI export rows: H=header, T=item, F=footer (Seema/Micropro or MARG ERP)."""
    t_rows = [row for row in rows if _clean_cell(_cell_at(row, 0)).upper() == "T"]
    if not t_rows:
        return None

    h_row = next((row for row in rows if _clean_cell(_cell_at(row, 0)).upper() == "H"), [])
    f_row = next((row for row in rows if _clean_cell(_cell_at(row, 0)).upper() == "F"), [])
    header_text = _rows_to_text([h_row] if h_row else [])
    edi_fmt = _detect_edi_htf_format(t_rows[0], header_text)

    issues: List[str] = []
    items: List[ImportedPurchaseItem] = []
    for row_no, row in enumerate(rows, start=1):
        if _clean_cell(_cell_at(row, 0)).upper() != "T":
            continue
        name = _clean_cell(_cell_at(row, 5))
        if not name:
            continue
        pack = _clean_cell(_cell_at(row, 6))
        expiry = _normalize_compact_expiry(_cell_at(row, 9)) or normalize_expiry(_cell_at(row, 9))
        rate, mrp, qty, free_qty, amount, gst_pct, hsn_code = _edi_htf_line_amounts(row, edi_fmt)
        mfg_col = 2 if edi_fmt == "marg" else 7
        if edi_fmt == "seema":
            line_cd_pct = _to_float(_cell_at(row, 20))
            line_pd_pct = _to_float(_cell_at(row, 21))
        elif edi_fmt == "seema_legacy":
            line_cd_pct = _to_float(_cell_at(row, 22))
            line_pd_pct = 0.0
        else:
            line_cd_pct = 0.0
            line_pd_pct = 0.0
        item = ImportedPurchaseItem(
            name=name,
            medicine_type=_type_from_record({
                "name": name,
                "pack": pack,
                "pkg_unit": pack,
            }),
            batch=_clean_cell(_cell_at(row, 8)),
            expiry=expiry,
            qty=qty,
            free_qty=free_qty,
            rate=rate,
            discount_pct=line_cd_pct,
            disc_column_value=line_pd_pct,
            disc_column_type="PERCENTAGE" if line_pd_pct else "ABSENT",
            gst_pct=gst_pct,
            mrp=mrp,
            hsn_code=hsn_code,
            manufacturer=_clean_cell(_cell_at(row, mfg_col)),
            pack=pack,
            amount=amount,
            source_row=row_no,
            raw={str(idx): _clean_cell(value) for idx, value in enumerate(row)},
        )
        item.validate()
        if item.issues:
            issues.append("Row {}: {}".format(row_no, "; ".join(item.issues)))
        items.append(item)

    if not items:
        return None

    text = _rows_to_text(rows[:40])
    details = extract_supplier_details(text)
    invoice_number = _clean_cell(_cell_at(h_row, 2)) or details.get("invoice_number", "")
    invoice_date = (
        _normalize_compact_date(_cell_at(h_row, 3))
        or _normalize_compact_date(_cell_at(h_row, 9))
        or details.get("invoice_date", "")
    )
    if edi_fmt == "seema":
        invoice_total = (
            _to_float(_cell_at(f_row, 23))
            or _to_float(_cell_at(f_row, 1))
            or _extract_total_amount(text, items)
        )
    elif edi_fmt == "seema_legacy":
        invoice_total = (
            _to_float(_cell_at(f_row, 21))
            or _to_float(_cell_at(f_row, 1))
            or _extract_total_amount(text, items)
        )
    elif edi_fmt == "marg":
        invoice_total = 0.0
    else:
        invoice_total = (
            _to_float(_cell_at(f_row, 23))
            or _to_float(_cell_at(f_row, 1))
            or _to_float(_cell_at(f_row, 21))
            or _extract_total_amount(text, items)
        )
    details["invoice_number"] = invoice_number
    details["invoice_date"] = invoice_date
    basename = os.path.basename(path).upper()
    if not details.get("supplier_name") and any(
        token in basename for token in ("SEEMA", "SJCR", "JCR00664")
    ):
        details["supplier_name"] = "SEEMA FRUITS PHARMA & VETERINARY DISTRIBUTORS"
    if not details.get("supplier_name") and "SWAMI_SAMARTH" in basename.replace(" ", "_"):
        details["supplier_name"] = "SHRI SWAMI SAMARTH MEDICAL AND AGENCY"

    parser_label = "EDI H/T/F ({})".format(edi_fmt)
    invoice = _invoice_from_details(
        details,
        merge_duplicate_items(items, issues),
        path,
        ext.lstrip("."),
        parser_label,
        invoice_total,
        issues,
        text,
    )
    _apply_edi_htf_footer(invoice, f_row, edi_fmt)
    return invoice


def _rows_from_dataframe(df: Any) -> List[List[Any]]:
    try:
        return df.where(df.notna(), "").values.tolist()
    except Exception:
        return []


def _records_from_table_rows(raw_rows: Sequence[Sequence[Any]], issues: List[str]) -> List[Dict[str, Any]]:
    rows = [_clean_row(row) for row in raw_rows if not _row_is_blank(row)]
    if not rows:
        return []

    header_idx, mapping = _find_header_map(rows)
    records: List[Dict[str, Any]] = []
    if mapping:
        for row_no, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            if _looks_like_header_row(row) or _looks_like_total_row(row):
                continue
            rec = {"source_row": row_no}
            for field_name, idx in mapping.items():
                rec[field_name] = _cell_at(row, idx)
            if _record_has_content(rec):
                records.append(rec)
        return records

    for row_no, row in enumerate(rows, start=1):
        rec = _positional_record(row)
        if rec:
            rec["source_row"] = row_no
            records.append(rec)
    return records


def _apply_ganesh_line_corrections(invoice: PurchaseInvoice) -> None:
    """Fix GST/HSN from Ganesh/vet text lines when table extraction misaligns columns."""
    text = invoice.raw_text or ""
    if not text:
        return
    parsed_lines = []
    for line in text.splitlines():
        rec = _parse_ganesh_vet_line(_clean_cell(line))
        if rec:
            parsed_lines.append(rec)
    if not parsed_lines:
        return
    for item in invoice.items:
        batch = _clean_key(item.batch)
        rate = round(_to_float(item.rate), 2)
        qty = round(_to_float(item.qty), 2)
        for rec in parsed_lines:
            if (
                _clean_key(rec.get("batch")) == batch
                and round(_to_float(rec.get("rate")), 2) == rate
                and round(_to_float(rec.get("qty")), 2) == qty
            ):
                item.gst_pct = _to_float(rec.get("gst_pct"))
                item.hsn_code = _clean_cell(rec.get("hsn_code")) or item.hsn_code
                item.manufacturer = _clean_cell(rec.get("manufacturer")) or item.manufacturer
                item.mrp = _to_float(rec.get("mrp")) or item.mrp
                item.amount = _to_float(rec.get("amount")) or item.amount
                break


def _records_from_text(text: str) -> List[Dict[str, Any]]:
    records = []
    for row_no, line in enumerate(text.splitlines(), start=1):
        line = _clean_cell(line)
        if not line or _looks_like_total_line(line):
            continue
        rec = _parse_ganesh_vet_line(line)
        if not rec:
            parts = _split_invoice_line(line)
            rec = _positional_record(parts)
        if rec:
            rec["source_row"] = row_no
            records.append(rec)
    return records


def _parse_ganesh_vet_line(line: str) -> Optional[Dict[str, Any]]:
    """
    JAI GANESH / veterinary format:
    [S.No] HSN MFG NAME... BATCH EXP MRP QTY FREE RATE AMOUNT DISC GST
    Example: 1 2309 ALEMB SHARKOFERROL VET 450G AN5569002 12/26 183.00 1.0 0.0 146.40 146.40 0.00 0
    """
    tokens = line.split()
    if len(tokens) < 14:
        return None
    start = 1 if tokens[0].isdigit() and len(tokens[0]) <= 3 else 0
    if start >= len(tokens) or not _looks_like_hsn(tokens[start]):
        return None
    try:
        gst_pct = _to_float(tokens[-1])
        _to_float(tokens[-2])
        amount = _to_float(tokens[-3])
        rate = _to_float(tokens[-4])
        free_qty = _to_float(tokens[-5])
        qty = _to_float(tokens[-6])
        mrp = _to_float(tokens[-7])
        expiry = normalize_expiry(tokens[-8])
        batch = tokens[-9]
    except (IndexError, ValueError):
        return None
    if qty <= 0 or rate <= 0:
        return None
    if not expiry or not batch:
        return None
    hsn = tokens[start]
    manufacturer = tokens[start + 1]
    name = " ".join(tokens[start + 2:-9]).strip()
    if not name:
        return None
    return {
        "hsn_code": hsn,
        "manufacturer": manufacturer,
        "name": name,
        "batch": batch,
        "expiry": expiry,
        "mrp": mrp,
        "qty": qty,
        "free_qty": free_qty,
        "rate": rate,
        "amount": amount or round(qty * rate, 2),
        "gst_pct": gst_pct,
        "discount_pct": 0,
    }


def _items_from_records(records: Sequence[Dict[str, Any]], issues: List[str]) -> List[ImportedPurchaseItem]:
    items: List[ImportedPurchaseItem] = []
    for rec in records:
        if not _record_has_content(rec):
            continue
        if _looks_like_non_item_name(rec.get("name")):
            continue
        item = _item_from_record(rec)
        if item.name or item.batch or item.qty or item.rate:
            apply_import_placeholders(item)
            item.validate()
            items.append(item)
            if item.issues:
                issues.append(
                    "Row {}: {}".format(
                        item.source_row or "?",
                        "; ".join(item.issues),
                    )
                )
    return items


def _merge_broken_records(records: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return merge_multiline_item_records(records)


def _item_from_record(rec: Dict[str, Any]) -> ImportedPurchaseItem:
    rec = apply_record_pricing(dict(rec))
    name = _clean_cell(rec.get("name"))
    pack = _clean_cell(rec.get("pack"))
    qty = _to_float(rec.get("qty"))
    free_qty = _to_float(rec.get("free_qty"))
    rate = _to_float(rec.get("rate"))
    amount = _to_float(rec.get("amount"))
    expiry = normalize_expiry(rec.get("expiry"))
    rec["qty_unit"] = _clean_cell(rec.get("qty_unit") or rec.get("unit"))
    rec["pkg_unit"] = _clean_cell(rec.get("pkg_unit") or pack)
    med_type = _type_from_record(rec)
    if _clean_cell(rec.get("medicine_type")):
        rec["medicine_type_source"] = "column"
    return ImportedPurchaseItem(
        name=name,
        medicine_type=med_type,
        batch=_clean_cell(rec.get("batch")),
        expiry=expiry,
        qty=qty,
        free_qty=free_qty,
        rate=rate,
        discount_pct=_to_float(rec.get("discount_pct")),
        disc_column_value=_to_float(rec.get("disc_column_value")),
        disc_column_type=str(rec.get("disc_column_type") or "ABSENT"),
        amount_validated=bool(rec.get("amount_validated", True)),
        gst_pct=_to_float(rec.get("gst_pct")),
        mrp=_to_float(rec.get("mrp")),
        hsn_code=_clean_cell(rec.get("hsn_code")),
        manufacturer=_clean_cell(rec.get("manufacturer")),
        pack=pack,
        amount=amount or round(qty * rate, 2),
        source_row=int(rec.get("source_row") or 0),
        raw=dict(rec),
    )


def _positional_record(row: Sequence[Any]) -> Dict[str, Any]:
    cells = [_clean_cell(cell) for cell in row if _clean_cell(cell)]
    if len(cells) == 1:
        cells = _split_invoice_line(cells[0])
    if not cells or _looks_like_header_row(cells):
        return {}
    if cells and _is_serial(cells[0]) and len(cells) > len(PDF_POSITIONAL_FIELDS):
        cells = cells[1:]
    if len(cells) < 8:
        return {}
    if not _looks_like_hsn(cells[0]) and len(cells) >= 12 and _looks_like_hsn(cells[1]):
        cells = cells[1:]

    if len(cells) >= 12:
        head = cells[:3]
        tail = cells[-8:]
        middle = cells[3:-8]
        if not middle:
            return {}
        values = list(head) + [" ".join(middle)] + tail
        return dict(zip(PDF_POSITIONAL_FIELDS, values))

    if len(cells) == 11:
        values = cells[:8] + [cells[8], "0", cells[9], cells[10]]
        return dict(zip(PDF_POSITIONAL_FIELDS, values))
    return {}


def _split_invoice_line(line: str) -> List[str]:
    line = _clean_cell(line)
    if "|" in line:
        return [_clean_cell(part) for part in line.split("|") if _clean_cell(part)]
    parts = [_clean_cell(part) for part in re.split(r"\t+|\s{2,}", line) if _clean_cell(part)]
    if len(parts) >= 8:
        return parts

    tokens = line.split()
    if len(tokens) >= 11 and _looks_like_hsn(tokens[0]):
        head = tokens[:3]
        tail = tokens[-8:]
        middle = tokens[3:-8]
        return head + [" ".join(middle)] + tail
    return parts


def _find_header_map(rows: Sequence[Sequence[Any]]) -> Tuple[int, Dict[str, int]]:
    best_idx = -1
    best_score = 0
    best_map: Dict[str, int] = {}
    for idx, row in enumerate(rows[:50]):
        mapping: Dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            field_name = _field_for_header(cell)
            if field_name and field_name not in mapping:
                mapping[field_name] = col_idx
        score = len(mapping)
        if "name" in mapping:
            score += 4
        if "qty" in mapping:
            score += 2
        if "rate" in mapping:
            score += 2
        if score > best_score:
            best_idx, best_score, best_map = idx, score, mapping
    if best_score >= 7 and "name" in best_map:
        return best_idx, best_map
    if "name" in best_map and (("qty" in best_map and "rate" in best_map) or "amount" in best_map):
        return best_idx, best_map
    return -1, {}


def _field_for_header(value: Any) -> str:
    return classify_header_field(value)


def _supplier_header_scope(text: str) -> str:
    upper = text.upper()
    for marker in ("M.R.P.", "MRP", "HSN CODE", "NAME OF PRODUCT", "SR. HSN"):
        idx = upper.find(marker)
        if idx > 0:
            text = text[:idx]
            break
    parts = re.split(
        r"Details Of Receiver|Billed To|\|\s*TO,|\bTo\s*:",
        text,
        maxsplit=1,
        flags=re.IGNORECASE,
    )
    return parts[0]


def _extract_gstin(text: str, supplier_side: bool = False) -> str:
    scope = _supplier_header_scope(text) if supplier_side else text
    pattern = r"\b([0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][A-Z0-9]Z[A-Z0-9])\b"
    m = re.search(r"GSTIN\s*:?\s*([0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][A-Z0-9]Z[A-Z0-9])", scope, flags=re.IGNORECASE)
    if m:
        return m.group(1).upper()
    m = re.search(pattern, scope, flags=re.IGNORECASE)
    return m.group(1).upper() if m else ""


def _extract_phone(text: str, supplier_side: bool = False) -> str:
    scope = _supplier_header_scope(text) if supplier_side else text
    patterns = (
        r"(?:PhoneNo|Contact No\.?|MOB(?:ile)?|Phone)\s*:?\s*(?:MOB:)?([0-9+\-\s,/]{10,30})",
        r"\b(\d{10})\b",
    )
    phones: List[str] = []
    for pattern in patterns:
        for match in re.finditer(pattern, scope, flags=re.IGNORECASE):
            digits = re.sub(r"\D", "", match.group(1))
            if len(digits) >= 10:
                phones.append(digits[-10:])
        if phones:
            break
    unique: List[str] = []
    for phone in phones:
        if phone not in unique:
            unique.append(phone)
    return ", ".join(unique[:2])


def _extract_dl_numbers(text: str, supplier_side: bool = False) -> str:
    scope = _supplier_header_scope(text) if supplier_side else text
    patterns = (
        r"(?:Lic\.?\s*No\.?|D\.\s*L\s*\.\s*No\.?)\s*:?\s*([^\n|]+?)(?=\s*(?:PhoneNo|Phone|Contact|GSTIN|Email|\||$))",
        r"DL\s*NO\.?\s*([^\n|]+?)(?=\s*(?:Inv\.|GSTIN|CONTACT|Date|Due Date|$))",
    )
    for pattern in patterns:
        m = re.search(pattern, scope, flags=re.IGNORECASE)
        if m:
            value = _clean_cell(m.group(1))
            value = re.sub(r"\s{2,}", " ", value)
            value = value.strip(" ,:-")
            value = re.sub(r"\s*,\s*Inv\..*$", "", value, flags=re.IGNORECASE).strip(" ,:-")
            if len(value) >= 4:
                return value
    return ""


def _pipe_table_header_rows(text: str) -> List[Tuple[str, List[str]]]:
    """Left-column text per piped header row (supplier block above item table)."""
    rows: List[Tuple[str, List[str]]] = []
    for line in text.splitlines()[:30]:
        if "|" not in line:
            continue
        cells = [_clean_cell(part) for part in line.split("|") if _clean_cell(part)]
        if not cells:
            continue
        left = cells[0]
        if left.upper() in ("TAX INVOICE", "H", "F"):
            continue
        if re.match(r"^[-=|\\/\s]+$", left):
            continue
        rows.append((left, cells))
    return rows


def _is_pipe_supplier_name_line(left: str) -> bool:
    low = left.lower()
    if any(token in low for token in ("hsn code", "name of product", "tax invoice")):
        return False
    return bool(
        re.search(
            r"(?:pharma|distributors?|veterinary|medical\s+(?:stores?|agency)|agencies?|chemist)",
            low,
            flags=re.IGNORECASE,
        )
    )


def _is_pipe_address_stop_line(left: str) -> bool:
    low = left.lower()
    return any(
        token in low
        for token in (
            "lic.no",
            "phoneno",
            "phone no",
            "gstin",
            "lr.no",
            "lr.date",
            "transport",
            "cases",
            "ewb no",
            "to,",
            "---",
            "hsn code",
        )
    )


def _is_pipe_address_line(left: str) -> bool:
    low = left.lower()
    if _is_pipe_address_stop_line(left):
        return False
    if any(token in low for token in ("invoice no", "order no", "credit memo", "invoice date", "order date")):
        return False
    if re.search(r"\border\s*no\b", low):
        return False
    if any(
        token in low
        for token in (
            "complex",
            "gala",
            "marg",
            "plot",
            "road",
            "ward",
            "shop",
            "floor",
            "malmatta",
            "cts no",
            "prop no",
            "niwane",
            "near lic",
            "umari",
            "shrawgi",
            "sant ",
        )
    ):
        return True
    if re.search(r"(?:\.|[-:])\s*\d{6}\b", left) or re.search(r"\bnandura\b", low, flags=re.IGNORECASE):
        return True
    if re.search(r"\b\d{6}\b", left) and re.search(r"[A-Za-z]{3,}", left):
        return True
    return False


def _prepend_supplier_city_to_address(address: str, supplier_name: str) -> str:
    m = re.search(r",\s*([A-Za-z][A-Za-z\s]{2,30})\s*$", supplier_name or "")
    if not m:
        return address
    city = _clean_cell(m.group(1))
    if not city or city.upper() in address.upper()[: len(city) + 4]:
        return address
    return f"{city} {address}".strip()


def _extract_pipe_supplier_address(text: str) -> str:
    """
    Two-column pipe GST invoices (Rathi, Seema Fruits, etc.):
    supplier name + address lines in the left column before Lic/Phone/GSTIN.
    """
    rows = _pipe_table_header_rows(text)
    if not rows:
        return ""

    supplier_idx = 0
    for idx, (left, _) in enumerate(rows):
        if _is_pipe_supplier_name_line(left):
            supplier_idx = idx
            break

    parts: List[str] = []
    for left, _cells in rows[supplier_idx + 1:]:
        if _is_pipe_address_stop_line(left):
            break
        if not _is_pipe_address_line(left):
            continue
        cleaned = left.strip(" ,")
        if cleaned:
            parts.append(cleaned)

    if not parts:
        return ""

    address = ", ".join(parts)
    supplier_left = rows[supplier_idx][0]
    return _prepend_supplier_city_to_address(address, supplier_left)


def _extract_supplier_address(text: str, supplier_name: str = "") -> str:
    if "|" in text:
        pipe_address = _extract_pipe_supplier_address(text)
        if pipe_address:
            return pipe_address

    scope = _supplier_header_scope(text)
    m = re.search(r"\b(\d{6})\b", scope)
    if m:
        pin = m.group(1)
        for line in scope.splitlines():
            if pin in line and not re.search(r"\bdl\s*no", line, flags=re.IGNORECASE):
                address = _clean_cell(line)
                if "|" in address:
                    left_cells = [
                        _clean_cell(part)
                        for part in line.split("|")
                        if _clean_cell(part)
                    ]
                    address = left_cells[0] if left_cells else ""
                if (
                    len(address) >= 8
                    and not _looks_like_dl_address_noise(address)
                    and not _looks_like_bad_supplier_address(address)
                ):
                    return address

    lines = [_clean_cell(line) for line in text.splitlines() if _clean_cell(line)]
    for idx, line in enumerate(lines[:25]):
        if _looks_like_bad_supplier_address(line):
            continue
        if supplier_name and supplier_name.lower()[:12] in line.lower():
            if re.search(r"\bfor\s*:", line, flags=re.IGNORECASE):
                continue
            collected: List[str] = []
            for follow in lines[idx + 1:idx + 5]:
                low = follow.lower()
                if _looks_like_dl_address_noise(follow):
                    continue
                if any(
                    token in low
                    for token in (
                        "gstin", "lic", "phone", "contact", "invoice",
                        "inv.no", "due date", "to :",
                    )
                ):
                    break
                if re.search(r"[A-Za-z]", follow):
                    cleaned = re.sub(
                        r"\s*State\s*Code\s*:.*$", "", follow, flags=re.IGNORECASE
                    ).strip(" ,")
                    if cleaned:
                        collected.append(cleaned)
            if collected:
                return ", ".join(collected)
    return ""


def _clean_pipe_cell(line: str) -> str:
    text = _clean_cell(line).strip("| ").strip()
    if set(text) <= set("-=|\\/"):
        return ""
    return text


def _extract_invoice_no_from_line(line: str) -> str:
    patterns = (
        r"\b(?:inv|invoice)\.?\s*no\.?\s*[:\-]?\s*([A-Z0-9][A-Z0-9/\- ]{1,25})",
        r"\b(?:invoice|inv)\s*(?:no\.?|number|#)\s*[:\-]?\s*([A-Z0-9][A-Z0-9/\- ]{1,25})",
        r"\binv\b\s*[:\-]?\s*([A-Z0-9][A-Z0-9/\-]{1,25})\b",
        r"\b(INV\b\s*[A-Z0-9/\-]+)\b",
        r"\b(BILL\s*(?:NO\.?|NUMBER)?\s*[:\-]?\s*[A-Z0-9/\-]+)\b",
        r"Invoice\s+No\.?\s*:?\s*(INV\s*[A-Z0-9/\-]+)",
    )
    for pattern in patterns:
        m = re.search(pattern, line, flags=re.IGNORECASE)
        if m:
            value = _clean_cell(m.group(1))
            value = re.split(r"\s{2,}|DATE|DT\.", value, flags=re.IGNORECASE)[0].strip()
            return value
    return ""


def _extract_supplier_from_line(line: str) -> str:
    m = re.match(
        r"^\s*(?:supplier\s+name|supplier|vendor\s+name|vendor|party\s+name|party)"
        r"\s*[:,-]?\s*(.+)$",
        line,
        flags=re.IGNORECASE,
    )
    if not m:
        return ""
    value = _clean_cell(m.group(1)).strip(" :-")
    return value if value and value.lower() not in ("name", "supplier") else ""


def _extract_date_from_line(line: str) -> str:
    if not re.search(r"date|dt\.?|invoice|inv|bill", line, flags=re.IGNORECASE):
        return ""
    m = re.search(r"(\d{1,2}[./-]\d{1,2}[./-]\d{2,4}|\d{4}[./-]\d{1,2}[./-]\d{1,2})", line)
    if not m:
        return ""
    return normalize_invoice_date(m.group(1))


def _extract_supplier_name(lines: Sequence[str]) -> str:
    blocked = (
        "tax invoice", "invoice", "gstin", "phone", "dl no", "d.l.no",
        "drug licence", "credit memo", "hsn code", "gst details",
        "no of items", "item total", "net payable", "lic.no",
    )
    keywords = (
        "pharma", "medical", "veterinary", "distributor", "distributors",
        "traders", "agency", "agencies", "store", "enterprise", "enterprises",
        "drugs", "chemist", "surgicals",
    )
    scored: List[Tuple[int, str]] = []
    for idx, line in enumerate(lines[:30]):
        candidate = _supplier_line_candidate(line, blocked)
        if not candidate:
            continue
        low = candidate.lower()
        if any(token in low for token in blocked):
            continue
        if len(candidate) < 4:
            continue
        score = 0
        if any(token in low for token in keywords):
            score += 10
        if candidate.upper() == candidate and re.search(r"[A-Z]", candidate):
            score += 3
        if "&" in candidate:
            score += 1
        score -= idx // 8
        if score > 0:
            previous = _supplier_line_candidate(lines[idx - 1], blocked) if idx > 0 else ""
            if previous and any(token in low for token in keywords):
                scored.append((score + 5, "{} {}".format(previous, candidate).strip(" :-")))
            scored.append((score, candidate.strip(" :-")))
    if scored:
        scored.sort(key=lambda item: item[0], reverse=True)
        return scored[0][1]
    for line in lines[:12]:
        candidate = _supplier_line_candidate(line, blocked)
        low = candidate.lower()
        if candidate and not any(token in low for token in blocked) and re.search(r"[A-Za-z]", candidate):
            return candidate.strip(" :-")
    return ""


def _supplier_line_candidate(line: str, blocked: Sequence[str]) -> str:
    text = _clean_cell(line).strip(" \t|:-")
    if not text:
        return ""
    if re.match(r"^[HTF]\b", text, flags=re.IGNORECASE):
        return ""
    if re.fullmatch(r"[-=|\\/\s\f]+", text):
        return ""
    if "|" in line:
        parts = [
            part.strip(" \t|:-")
            for part in line.split("|")
            if part.strip(" \t|:-")
        ]
        for part in parts:
            low = part.lower()
            if any(token in low for token in blocked):
                continue
            if re.search(r"[A-Za-z]", part) and not re.search(r"\d{3,}", part):
                return part
        return ""
    low = text.lower()
    if any(token in low for token in blocked):
        return ""
    if not re.search(r"[A-Za-z]", text):
        return ""
    return text


def _extract_bill_item_count(text: str) -> int:
    """Read 'No. of items 18' from Tuljai-style invoice footer."""
    m = re.search(r"No\.?\s*of\s*items?\s*(\d+)", text or "", flags=re.IGNORECASE)
    if m:
        try:
            return int(m.group(1))
        except ValueError:
            pass
    return 0


def _extract_total_amount(text: str, items: Sequence[ImportedPurchaseItem]) -> float:
    """Prefer net payable from footer; fall back to sum of line amounts."""
    from core.purchase_invoice_engine import extract_invoice_totals

    totals = extract_invoice_totals(text or "")
    if totals.get("net_payable"):
        return round(float(totals["net_payable"]), 2)

    candidates: List[float] = []
    patterns = (
        r"net\s*amt\s*r/?o\D+([\d,]+(?:\.\d{1,2})?)",
        r"(?:net\s+payable|grand\s+total|net\s+amount|invoice\s+total)\D+([\d,]+(?:\.\d{1,2})?)",
    )
    for line in reversed(text.splitlines()):
        for pattern in patterns:
            m = re.search(pattern, line, flags=re.IGNORECASE)
            if m:
                amount = _to_float(m.group(1))
                if amount:
                    candidates.append(amount)
    if candidates:
        return round(candidates[0], 2)
    return round(sum(float(i.amount or 0) for i in items), 2)


def _add_quantity_metadata(item: Dict[str, Any], pack: Any) -> None:
    med_type = str(item.get("type", ""))
    if is_strip_count_type(med_type):
        tps = _extract_tablets_per_strip(pack)
        item["tablets_per_stripe"] = tps
        item["total_tablets"] = float(item.get("qty", 0) or 0) * tps
        item["free_tablets"] = float(item.get("free_qty", 0) or 0) * tps
        return
    item["quantity_value"] = _clean_cell(pack) or "1"
    layout_unit = get_type_measure_unit(med_type)
    item["auto_unit"] = layout_unit or TYPE_UNITS.get(med_type.lower(), "")


def _extract_tablets_per_strip(pack: Any) -> int:
    text = _clean_cell(pack)
    matches = re.findall(r"\d+", text)
    if matches:
        try:
            return max(1, int(matches[0]))
        except ValueError:
            pass
    return 1


def _merge_purchase_page_items(items: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    merged: Dict[Tuple[Any, ...], Dict[str, Any]] = {}
    output: List[Dict[str, Any]] = []
    for item in items:
        key = (
            _clean_key(item.get("name")),
            _clean_key(item.get("batch")),
            item.get("expiry"),
            round(float(item.get("rate", 0) or 0), 4),
            round(float(item.get("gst_pct", 0) or 0), 4),
            round(float(item.get("mrp", 0) or 0), 4),
            _clean_key(item.get("hsn_code")),
            _clean_key(item.get("manufacturer")),
        )
        if key not in merged:
            merged[key] = item
            output.append(item)
            continue
        target = merged[key]
        target["qty"] = round(float(target.get("qty", 0) or 0) + float(item.get("qty", 0) or 0), 4)
        target["free_qty"] = round(
            float(target.get("free_qty", 0) or 0) + float(item.get("free_qty", 0) or 0),
            4,
        )
        _add_quantity_metadata(target, target.get("quantity_value", ""))
    return output


def _apply_invoice_discounts(purchase_page: Any, invoice: PurchaseInvoice) -> None:
    """Apply product + cash discount from invoice footer to overall discount field."""
    total_disc = round(
        float(getattr(invoice, "product_discount", 0) or 0)
        + float(getattr(invoice, "cash_discount", 0) or 0),
        2,
    )
    if total_disc <= 0:
        return
    if hasattr(purchase_page, "overall_discount"):
        _set_entry(purchase_page.overall_discount, "{:.2f}".format(total_disc))
    cash_pct = float(getattr(invoice, "cash_discount_pct", 0) or 0)
    if cash_pct and hasattr(purchase_page, "overall_discount_pct"):
        _set_entry(purchase_page.overall_discount_pct, "{:.2f}".format(cash_pct))
    elif hasattr(purchase_page, "sync_overall_discount_fields"):
        purchase_page.sync_overall_discount_fields('rupees')
    purchase_page.calculate_total()


def _apply_invoice_rounding(purchase_page: Any, invoice_total: Any) -> None:
    if not invoice_total or not hasattr(purchase_page, "rounding_entry"):
        return
    calc = getattr(purchase_page, "_last_calc", None) or {}
    calculated_total = float(calc.get("total_amount", 0) or 0)
    if not calculated_total:
        return
    diff = round(float(invoice_total or 0) - calculated_total, 2)
    if abs(diff) > 5:
        return
    _set_entry(purchase_page.rounding_entry, "{:.2f}".format(diff))
    purchase_page.calculate_total()


def _match_type(med_type: str, available_types: Iterable[str]) -> str:
    wanted = (med_type or "").strip()
    for option in available_types or []:
        if str(option).lower() == wanted.lower():
            return str(option)
    return wanted or "Tablet"


def _set_entry(widget: Any, value: Any) -> None:
    widget.delete(0, "end")
    widget.insert(0, "" if value is None else str(value))


def _valid_date(year: int, month: int, day: int) -> str:
    try:
        return datetime(year, month, day).strftime("%Y-%m-%d")
    except ValueError:
        return ""


def _format_mmyy(month_raw: Any, year_raw: Any) -> str:
    try:
        month = int(str(month_raw).strip())
        year_text = str(year_raw).strip()
        year = int(year_text)
        if len(year_text) == 2:
            year += 2000
        if 1 <= month <= 12 and 2000 <= year <= 2099:
            return "{:02d}/{:02d}".format(month, year % 100)
    except Exception:
        pass
    return ""


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return round(float(value), 4)
        except Exception:
            return 0.0
    text = _clean_cell(value)
    if not text:
        return 0.0
    text = text.replace(",", "")
    text = re.sub(r"(rs\.?|inr|rs|/-|%)", "", text, flags=re.IGNORECASE)
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in ("", "-", ".", "-."):
        return 0.0
    try:
        return round(float(text), 4)
    except ValueError:
        return 0.0


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    try:
        if value != value:
            return ""
    except Exception:
        pass
    text = str(value).replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return "" if text.lower() in ("none", "nan", "null") else text


def _clean_text(text: str) -> str:
    return text.replace("\r\n", "\n").replace("\r", "\n")


def _clean_row(row: Sequence[Any]) -> List[str]:
    return [_clean_cell(cell) for cell in row]


def _row_is_blank(row: Sequence[Any]) -> bool:
    return not any(_clean_cell(cell) for cell in row)


def _cell_at(row: Sequence[Any], idx: int) -> Any:
    try:
        return row[idx]
    except Exception:
        return ""


def _normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _clean_cell(value).lower())


def _clean_key(value: Any) -> str:
    return re.sub(r"\s+", " ", _clean_cell(value).upper()).strip()


def _looks_like_header_row(row: Sequence[Any]) -> bool:
    mapped = [_field_for_header(cell) for cell in row]
    return len([field for field in mapped if field]) >= 3


def _looks_like_total_row(row: Sequence[Any]) -> bool:
    if is_skip_row(row):
        return True
    return _looks_like_total_line(" ".join(_clean_cell(cell) for cell in row))


def _looks_like_total_line(line: str) -> bool:
    low = line.lower()
    return any(
        token in low
        for token in (
            "grand total", "net amount", "net payable", "round off", "taxable total",
            "subtotal", "gross amount", "total c/f", "total b/f", "amount in words",
            "item total", "prod discount", "cash discount",
        )
    )


def _record_has_content(rec: Dict[str, Any]) -> bool:
    return any(_clean_cell(rec.get(key)) for key in ("name", "batch", "expiry", "qty", "rate", "amount"))


def _looks_like_hsn(value: Any) -> bool:
    text = _clean_cell(value)
    return bool(re.fullmatch(r"\d{4,10}", text))


def _is_serial(value: Any) -> bool:
    text = _clean_cell(value)
    return bool(re.fullmatch(r"\d{1,3}", text))


def _rows_to_text(rows: Sequence[Sequence[Any]]) -> str:
    return "\n".join(
        " ".join(_clean_cell(cell) for cell in row if _clean_cell(cell))
        for row in rows
        if not _row_is_blank(row)
    )


def _ensure_file(path: str) -> None:
    if not path or not os.path.exists(path):
        raise InvoiceParseError("File not found: {}".format(path))
    if not os.path.isfile(path):
        raise InvoiceParseError("Not a file: {}".format(path))


def _config_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.join(os.environ.get("LOCALAPPDATA", os.path.expanduser("~")), "VeterinaryApp")
    return os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "config")
