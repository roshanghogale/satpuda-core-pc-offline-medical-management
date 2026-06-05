"""
Mode-aware master medicine database service.

Medical mode:
  - keeps a populated SQLite master DB built from bundled Excel.
Veterinary mode:
  - keeps master DB cleared/removed.
"""
from __future__ import annotations

import os
import sqlite3
import sys
from typing import Callable, Dict, List, Optional, Tuple

ProgressCb = Optional[Callable[[int, str], None]]

MASTER_TABLE = "medicines_master"


def _config_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.join(
            os.environ.get("LOCALAPPDATA", os.path.expanduser("~")),
            "VeterinaryApp",
        )
    return os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "config",
    )


def get_master_db_path() -> str:
    return os.path.join(_config_dir(), "master_medicine.db")


def _bundled_excel_path() -> str:
    if getattr(sys, "frozen", False):
        return os.path.join(sys._MEIPASS, "assets", "medicines_master_with_cdsco.xlsx")
    return os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "assets",
        "medicines_master_with_cdsco.xlsx",
    )


def _connect_master() -> sqlite3.Connection:
    os.makedirs(_config_dir(), exist_ok=True)
    conn = sqlite3.connect(get_master_db_path())
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    return conn


def _init_schema(conn: sqlite3.Connection) -> None:
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {MASTER_TABLE} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL COLLATE NOCASE UNIQUE,
            manufacturer TEXT,
            mrp REAL,
            content_drug TEXT,
            med_type TEXT,
            pack_size TEXT
        )
        """
    )
    conn.execute(
        f"CREATE INDEX IF NOT EXISTS idx_{MASTER_TABLE}_name ON {MASTER_TABLE}(name COLLATE NOCASE)"
    )
    conn.commit()


def clear_master_db() -> None:
    path = get_master_db_path()
    if os.path.exists(path):
        try:
            os.remove(path)
        except OSError:
            # If file is locked, fallback to truncating table.
            conn = _connect_master()
            try:
                _init_schema(conn)
                conn.execute(f"DELETE FROM {MASTER_TABLE}")
                conn.commit()
            finally:
                conn.close()


def _count_rows(conn: sqlite3.Connection) -> int:
    _init_schema(conn)
    cur = conn.cursor()
    cur.execute(f"SELECT COUNT(*) FROM {MASTER_TABLE}")
    return int(cur.fetchone()[0] or 0)


def _detect_type(name: str, form: str) -> str:
    text = f"{name} {form}".lower()
    rules = [
        ("tablet", "Tablet"),
        ("capsule", "Capsule"),
        ("syrup", "Syrup"),
        ("injection", "Injection"),
        ("ointment", "Ointment"),
        ("cream", "Ointment"),
        ("gel", "Gel"),
        ("drop", "Drops"),
        ("powder", "Powder"),
        ("bolus", "Bolus"),
        ("vaccine", "Vaccine"),
        ("liniment", "Liniment"),
        ("granule", "Granules"),
        ("liquid", "Liquid"),
    ]
    for kw, med_type in rules:
        if kw in text:
            return med_type
    return "Others"


def ensure_master_db_ready(progress_cb: ProgressCb = None) -> Tuple[bool, int, str]:
    """
    Ensure master DB exists and is populated.
    Returns (ready, row_count, message).
    """
    try:
        conn = _connect_master()
        try:
            existing = _count_rows(conn)
            if existing > 0:
                if progress_cb:
                    progress_cb(100, f"Master medicines ready ({existing:,})")
                return True, existing, "already-populated"
        finally:
            conn.close()

        excel_path = _bundled_excel_path()
        if not os.path.exists(excel_path):
            if progress_cb:
                progress_cb(100, "Master Excel not found")
            return False, 0, "excel-not-found"

        if progress_cb:
            progress_cb(5, "Reading master medicines Excel...")

        import re
        import openpyxl

        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        max_rows = int(ws.max_row or 1)
        processed = 0

        conn = _connect_master()
        try:
            _init_schema(conn)
            conn.execute(f"DELETE FROM {MASTER_TABLE}")
            conn.commit()

            batch: List[Tuple[str, str, float, str, str, str]] = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                processed += 1
                name = str(row[0]).strip() if row and row[0] else ""
                if not name or name.lower() == "none":
                    if processed % 2000 == 0 and progress_cb:
                        pct = min(98, 5 + int((processed / max(max_rows, 1)) * 90))
                        progress_cb(pct, f"Loading medicines... {processed:,}/{max_rows:,}")
                    continue

                mfg = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                try:
                    mrp = float(row[10]) if len(row) > 10 and row[10] is not None else 0.0
                except Exception:
                    mrp = 0.0
                salt = str(row[11]).strip() if len(row) > 11 and row[11] else ""
                salt = re.sub(r"\s*\+\s*nan\s*", "", salt).strip().strip("+").strip()
                form = str(row[16]).strip() if len(row) > 16 and row[16] else ""
                pack = str(row[17]).strip() if len(row) > 17 and row[17] else ""
                batch.append((name, mfg, mrp, salt, _detect_type(name, form), pack))

                if len(batch) >= 5000:
                    conn.executemany(
                        f"""
                        INSERT INTO {MASTER_TABLE}
                        (name, manufacturer, mrp, content_drug, med_type, pack_size)
                        VALUES (?, ?, ?, ?, ?, ?)
                        ON CONFLICT(name) DO UPDATE SET
                            manufacturer=excluded.manufacturer,
                            mrp=excluded.mrp,
                            content_drug=excluded.content_drug,
                            med_type=excluded.med_type,
                            pack_size=excluded.pack_size
                        """,
                        batch,
                    )
                    conn.commit()
                    batch.clear()

                if processed % 2000 == 0 and progress_cb:
                    pct = min(98, 5 + int((processed / max(max_rows, 1)) * 90))
                    progress_cb(pct, f"Loading medicines... {processed:,}/{max_rows:,}")

            if batch:
                conn.executemany(
                    f"""
                    INSERT INTO {MASTER_TABLE}
                    (name, manufacturer, mrp, content_drug, med_type, pack_size)
                    VALUES (?, ?, ?, ?, ?, ?)
                    ON CONFLICT(name) DO UPDATE SET
                        manufacturer=excluded.manufacturer,
                        mrp=excluded.mrp,
                        content_drug=excluded.content_drug,
                        med_type=excluded.med_type,
                        pack_size=excluded.pack_size
                    """,
                    batch,
                )
                conn.commit()

            count = _count_rows(conn)
            if progress_cb:
                progress_cb(100, f"Master medicines loaded ({count:,})")
            return True, count, "imported"
        finally:
            conn.close()
            wb.close()
    except Exception as exc:
        if progress_cb:
            progress_cb(100, f"Master load failed: {exc}")
        return False, 0, str(exc)


def ensure_mode_master_state(mode: str, progress_cb: ProgressCb = None) -> Tuple[bool, int, str]:
    mode = (mode or "medical").strip().lower()
    if mode == "veterinary":
        clear_master_db()
        if progress_cb:
            progress_cb(100, "Veterinary mode: master DB cleared")
        return True, 0, "cleared-for-veterinary"
    return ensure_master_db_ready(progress_cb=progress_cb)


def search_master_names(typed: str, limit: int = 50) -> List[str]:
    query = (typed or "").strip()
    conn = _connect_master()
    try:
        _init_schema(conn)
        cur = conn.cursor()
        if not query:
            cur.execute(
                f"SELECT name FROM {MASTER_TABLE} ORDER BY name COLLATE NOCASE LIMIT ?",
                (limit,),
            )
            return [r[0] for r in cur.fetchall()]

        cur.execute(
            f"""
            SELECT name FROM {MASTER_TABLE}
            WHERE name LIKE ? COLLATE NOCASE
            ORDER BY name COLLATE NOCASE
            LIMIT ?
            """,
            (f"{query}%", limit),
        )
        prefix = [r[0] for r in cur.fetchall()]

        if len(prefix) >= limit:
            return prefix[:limit]

        seen = {n.lower() for n in prefix}
        cur.execute(
            f"""
            SELECT name FROM {MASTER_TABLE}
            WHERE name LIKE ? COLLATE NOCASE
              AND name NOT LIKE ? COLLATE NOCASE
            ORDER BY name COLLATE NOCASE
            LIMIT ?
            """,
            (f"%{query}%", f"{query}%", limit - len(prefix)),
        )
        contains = [r[0] for r in cur.fetchall() if r[0].lower() not in seen]
        return (prefix + contains)[:limit]
    finally:
        conn.close()


def get_all_master_names(limit: int = 0) -> List[str]:
    conn = _connect_master()
    try:
        _init_schema(conn)
        cur = conn.cursor()
        if limit and int(limit) > 0:
            cur.execute(
                f"SELECT name FROM {MASTER_TABLE} ORDER BY name COLLATE NOCASE LIMIT ?",
                (int(limit),),
            )
        else:
            cur.execute(f"SELECT name FROM {MASTER_TABLE} ORDER BY name COLLATE NOCASE")
        return [r[0] for r in cur.fetchall()]
    finally:
        conn.close()


def lookup_master_details(name: str) -> Dict[str, object]:
    medicine_name = (name or "").strip()
    if not medicine_name:
        return {}
    conn = _connect_master()
    try:
        _init_schema(conn)
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT med_type, manufacturer, mrp, content_drug, pack_size
            FROM {MASTER_TABLE}
            WHERE name = ? COLLATE NOCASE
            LIMIT 1
            """,
            (medicine_name,),
        )
        row = cur.fetchone()
        if not row:
            return {}
        return {
            "type": row[0] or "",
            "manufacturer": row[1] or "",
            "mrp": float(row[2] or 0),
            "content_drug": row[3] or "",
            "pack_size": row[4] or "",
        }
    finally:
        conn.close()


def upsert_master_medicine(
    name: str,
    manufacturer: str = "",
    mrp: float = 0.0,
    content_drug: str = "",
    med_type: str = "",
    pack_size: str = "",
) -> None:
    medicine_name = (name or "").strip()
    if not medicine_name:
        return
    conn = _connect_master()
    try:
        _init_schema(conn)
        conn.execute(
            f"""
            INSERT INTO {MASTER_TABLE}
            (name, manufacturer, mrp, content_drug, med_type, pack_size)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(name) DO UPDATE SET
                manufacturer=excluded.manufacturer,
                mrp=excluded.mrp,
                content_drug=excluded.content_drug,
                med_type=excluded.med_type,
                pack_size=excluded.pack_size
            """,
            (
                medicine_name,
                (manufacturer or "").strip(),
                float(mrp or 0),
                (content_drug or "").strip(),
                (med_type or "").strip(),
                (pack_size or "").strip(),
            ),
        )
        conn.commit()
    finally:
        conn.close()


def sync_master_with_inventory(inventory_conn: sqlite3.Connection) -> int:
    """
    Upsert medicines from local inventory DB into master DB.
    Returns number of medicines processed.
    """
    if inventory_conn is None:
        return 0
    try:
        cur = inventory_conn.cursor()
        cur.execute(
            """
            SELECT m.name,
                   COALESCE(m.manufacturer, ''),
                   COALESCE(m.mrp, 0),
                   COALESCE(m.content_drug, ''),
                   COALESCE(m.type, ''),
                   COALESCE(m.unit, '')
            FROM medicines m
            JOIN (
                SELECT name, MAX(id) AS max_id
                FROM medicines
                GROUP BY name
            ) latest ON latest.max_id = m.id
            WHERE COALESCE(m.name, '') <> ''
            """
        )
        rows = cur.fetchall()
    except Exception:
        return 0

    if not rows:
        return 0

    conn = _connect_master()
    try:
        _init_schema(conn)
        conn.executemany(
            f"""
            INSERT INTO {MASTER_TABLE}
            (name, manufacturer, mrp, content_drug, med_type, pack_size)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(name) DO UPDATE SET
                manufacturer=excluded.manufacturer,
                mrp=excluded.mrp,
                content_drug=excluded.content_drug,
                med_type=excluded.med_type,
                pack_size=excluded.pack_size
            """,
            [
                (
                    str(r[0] or "").strip(),
                    str(r[1] or "").strip(),
                    float(r[2] or 0),
                    str(r[3] or "").strip(),
                    str(r[4] or "").strip(),
                    str(r[5] or "").strip(),
                )
                for r in rows
                if str(r[0] or "").strip()
            ],
        )
        conn.commit()
        return len(rows)
    finally:
        conn.close()


def sync_master_with_inventory_db_path(inventory_db_path: str) -> int:
    """
    Thread-safe variant: opens its own inventory DB connection by path.
    """
    if not inventory_db_path:
        return 0
    conn = sqlite3.connect(inventory_db_path)
    try:
        return sync_master_with_inventory(conn)
    finally:
        conn.close()
