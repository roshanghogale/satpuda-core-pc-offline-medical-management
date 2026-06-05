"""
export_manager.py
-----------------
Centralised export engine for CSV, Excel and PDF (HTML-based).
All exports open a save-file dialog then write the chosen format.

Public API
----------
export_data(parent, title, headers, rows, default_name)
    -> shows format chooser dialog then saves

export_all_combined(parent, sections)
    -> sections = list of (title, headers, rows)
    -> one dialog, one file with all sections combined
"""

import os
import csv
import tempfile
import tkinter as tk
from tkinter import ttk, filedialog, messagebox


def _apply_icon(window):
    """Apply satpuda_logo to any Tk/Toplevel window."""
    try:
        from core.window_icon import apply_window_icon
        apply_window_icon(window)
    except Exception:
        pass


# ── Format chooser dialog ─────────────────────────────────────────────────────

def export_data(parent, title, headers, rows, default_name='export'):
    """Show format picker then save. rows = list of tuples/lists."""
    from core.scroll_manager import open_dialog

    top = parent.winfo_toplevel()
    dlg = open_dialog(top, f"Export — {title}", width=360, height=200, resizable=False)
    body = dlg.content

    ttk.Label(body, text=f"Export: {title}",
              font=('Segoe UI', 11, 'bold')).pack(pady=(12, 6), padx=12)
    ttk.Label(body, text="Choose format:").pack(padx=12)

    fmt_var = tk.StringVar(value='csv')
    btn_row = ttk.Frame(body)
    btn_row.pack(pady=10, padx=12)
    for text, val in [('CSV', 'csv'), ('Excel (.xlsx)', 'xlsx'), ('PDF (HTML)', 'pdf')]:
        ttk.Radiobutton(btn_row, text=text, variable=fmt_var,
                        value=val).pack(side=tk.LEFT, padx=8)

    def _do_export():
        fmt = fmt_var.get()
        dlg.destroy()
        if fmt == 'csv':
            _save_csv(top, headers, rows, default_name)
        elif fmt == 'xlsx':
            _save_xlsx(top, headers, rows, default_name)
        else:
            _save_pdf(top, title, headers, rows, default_name)

    ttk.Button(dlg.footer, text="Export", command=_do_export).pack(side=tk.LEFT, padx=6)
    ttk.Button(dlg.footer, text="Cancel", command=dlg.destroy).pack(side=tk.LEFT, padx=6)


# ── Export All Combined ───────────────────────────────────────────────────────

def export_all_combined(parent, sections):
    """One dialog, one file combining all sections.
    sections = list of (title, headers, rows)
    CSV: sections separated by blank line + title header.
    Excel: one sheet per section.
    HTML: one page with all sections.
    """
    from core.scroll_manager import open_dialog

    top = parent.winfo_toplevel()
    dlg = open_dialog(top, "Export All Data", width=380, height=220, resizable=False)
    body = dlg.content

    ttk.Label(body, text="Export All Data",
              font=('Segoe UI', 12, 'bold')).pack(pady=(12, 4), padx=12)
    total = sum(len(r) for _, _, r in sections)
    ttk.Label(body, text=f"{len(sections)} sections  •  {total} total records",
              foreground='gray').pack(pady=(0, 8), padx=12)
    ttk.Label(body, text="Choose format:").pack(padx=12)

    fmt_var = tk.StringVar(value='xlsx')
    btn_row = ttk.Frame(body)
    btn_row.pack(pady=8, padx=12)
    for text, val in [('CSV', 'csv'), ('Excel (.xlsx)', 'xlsx'), ('PDF (HTML)', 'pdf')]:
        ttk.Radiobutton(btn_row, text=text, variable=fmt_var,
                        value=val).pack(side=tk.LEFT, padx=8)

    def _do_export():
        fmt = fmt_var.get()
        dlg.destroy()
        if fmt == 'csv':
            _save_all_csv(top, sections)
        elif fmt == 'xlsx':
            _save_all_xlsx(top, sections)
        else:
            _save_all_pdf(top, sections)

    ttk.Button(dlg.footer, text="Export", command=_do_export).pack(side=tk.LEFT, padx=6)
    ttk.Button(dlg.footer, text="Cancel", command=dlg.destroy).pack(side=tk.LEFT, padx=6)


def _save_all_csv(parent, sections):
    path = filedialog.asksaveasfilename(
        parent=parent,
        defaultextension='.csv',
        filetypes=[('CSV files', '*.csv')],
        initialfile='export_all.csv')
    if not path:
        return
    try:
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f)
            for i, (title, headers, rows) in enumerate(sections):
                if i > 0:
                    w.writerow([])  # blank separator
                w.writerow([f'=== {title} ==='])
                w.writerow(headers)
                w.writerows(rows)
        messagebox.showinfo("Exported", f"Saved to:\n{path}", parent=parent)
    except Exception as e:
        messagebox.showerror("Export Error", str(e), parent=parent)


def _save_all_xlsx(parent, sections):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        messagebox.showwarning(
            "openpyxl not installed",
            "openpyxl is required for Excel export.\n"
            "Install it with:  pip install openpyxl\n\n"
            "Saving as CSV instead.",
            parent=parent)
        _save_all_csv(parent, sections)
        return

    path = filedialog.asksaveasfilename(
        parent=parent,
        defaultextension='.xlsx',
        filetypes=[('Excel files', '*.xlsx')],
        initialfile='export_all.xlsx')
    if not path:
        return
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        hdr_fill = PatternFill('solid', fgColor='2C3E50')
        hdr_font = Font(color='FFFFFF', bold=True)
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for title, headers, rows in sections:
            ws = wb.create_sheet(title=title[:31])
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            for ri, row in enumerate(rows, 2):
                bg = 'F7F7F7' if ri % 2 == 0 else 'FFFFFF'
                fill = PatternFill('solid', fgColor=bg)
                for ci, val in enumerate(row, 1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left')
            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        wb.save(path)
        messagebox.showinfo("Exported", f"Saved to:\n{path}", parent=parent)
    except Exception as e:
        messagebox.showerror("Export Error", str(e), parent=parent)


def _save_all_pdf(parent, sections):
    path = filedialog.asksaveasfilename(
        parent=parent,
        defaultextension='.html',
        filetypes=[('HTML/PDF files', '*.html')],
        initialfile='export_all.html')
    if not path:
        return
    try:
        from datetime import datetime
        date_str = datetime.now().strftime('%d/%m/%Y %H:%M')

        def esc(v):
            return str(v).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

        body = ''
        for title, headers, rows in sections:
            hdr_html = ''.join(f'<th>{esc(h)}</th>' for h in headers)
            rows_html = ''.join(
                f'<tr class="{"even" if i%2==0 else "odd"}">'
                + ''.join(f'<td>{esc(v)}</td>' for v in row)
                + '</tr>'
                for i, row in enumerate(rows)
            )
            body += f"""
<h3>{esc(title)}</h3>
<p class="meta">{len(rows)} records</p>
<table><thead><tr>{hdr_html}</tr></thead><tbody>{rows_html}</tbody></table>
<div style="margin-bottom:12mm"></div>
"""

        html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>Export All</title>
<style>
  @page {{ size: A4 landscape; margin: 10mm; }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; font-size: 8.5pt; color: #000; }}
  h2 {{ text-align:center; font-size:13pt; margin-bottom:2mm; }}
  h3 {{ font-size:11pt; margin:6mm 0 2mm; color:#2c3e50; border-bottom:1pt solid #2c3e50; padding-bottom:1mm; }}
  .meta {{ font-size:8pt; color:#555; margin-bottom:2mm; }}
  table {{ width:100%; border-collapse:collapse; font-size:8pt; margin-bottom:4mm; }}
  thead th {{ background:#2c3e50; color:#fff; padding:1.5mm 2mm; text-align:left; border:0.3pt solid #000; }}
  tbody td {{ padding:1.2mm 2mm; border:0.3pt solid #ccc; }}
  tr.even {{ background:#f7f7f7; }} tr.odd {{ background:#fff; }}
  .print-btn {{ display:block; margin:5mm auto; padding:2mm 10mm; font-size:11pt;
                background:#2c3e50; color:white; border:none; border-radius:3px; cursor:pointer; }}
  @media print {{ .print-btn {{ display:none; }} }}
</style></head><body>
<h2>Full Data Export</h2>
<p class="meta" style="text-align:center">Generated: {date_str}</p>
{body}
<button class="print-btn" onclick="window.print()">&#128424; Print / Save as PDF</button>
</body></html>"""

        with open(path, 'w', encoding='utf-8') as f:
            f.write(html)
        if os.name == 'nt':
            os.startfile(path)
        else:
            import subprocess
            subprocess.Popen(['xdg-open', path])
        messagebox.showinfo("Exported",
            f"Opened in browser.\nUse Ctrl+P → Save as PDF.\n\nFile: {path}",
            parent=parent)
    except Exception as e:
        messagebox.showerror("Export Error", str(e), parent=parent)


# ── CSV ───────────────────────────────────────────────────────────────────────

def _save_csv(parent, headers, rows, default_name):
    path = filedialog.asksaveasfilename(
        parent=parent,
        defaultextension='.csv',
        filetypes=[('CSV files', '*.csv')],
        initialfile=f'{default_name}.csv')
    if not path:
        return
    try:
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f)
            w.writerow(headers)
            w.writerows(rows)
        messagebox.showinfo("Exported", f"Saved to:\n{path}", parent=parent)
    except Exception as e:
        messagebox.showerror("Export Error", str(e), parent=parent)


# ── Excel ─────────────────────────────────────────────────────────────────────

def _save_xlsx(parent, headers, rows, default_name):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        # Fallback: save as CSV with .xlsx extension hint
        messagebox.showwarning(
            "openpyxl not installed",
            "openpyxl is required for Excel export.\n"
            "Install it with:  pip install openpyxl\n\n"
            "Saving as CSV instead.",
            parent=parent)
        _save_csv(parent, headers, rows, default_name)
        return

    path = filedialog.asksaveasfilename(
        parent=parent,
        defaultextension='.xlsx',
        filetypes=[('Excel files', '*.xlsx')],
        initialfile=f'{default_name}.xlsx')
    if not path:
        return
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = default_name[:31]

        # Header style
        hdr_fill = PatternFill('solid', fgColor='2C3E50')
        hdr_font = Font(color='FFFFFF', bold=True)
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Data rows
        for ri, row in enumerate(rows, 2):
            bg = 'F7F7F7' if ri % 2 == 0 else 'FFFFFF'
            fill = PatternFill('solid', fgColor=bg)
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(horizontal='left')

        # Auto column width
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        wb.save(path)
        messagebox.showinfo("Exported", f"Saved to:\n{path}", parent=parent)
    except Exception as e:
        messagebox.showerror("Export Error", str(e), parent=parent)


# ── PDF (HTML) ────────────────────────────────────────────────────────────────

def _save_pdf(parent, title, headers, rows, default_name):
    path = filedialog.asksaveasfilename(
        parent=parent,
        defaultextension='.html',
        filetypes=[('HTML/PDF files', '*.html')],
        initialfile=f'{default_name}.html')
    if not path:
        return
    try:
        html = _build_html(title, headers, rows)
        with open(path, 'w', encoding='utf-8') as f:
            f.write(html)
        # Open in browser for printing
        if os.name == 'nt':
            os.startfile(path)
        else:
            import subprocess
            subprocess.Popen(['xdg-open', path])
        messagebox.showinfo(
            "Exported",
            f"Opened in browser.\nUse Ctrl+P → Save as PDF.\n\nFile: {path}",
            parent=parent)
    except Exception as e:
        messagebox.showerror("Export Error", str(e), parent=parent)


def _build_html(title, headers, rows):
    from datetime import datetime
    date_str = datetime.now().strftime('%d/%m/%Y %H:%M')

    def esc(v):
        return str(v).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    header_html = ''.join(f'<th>{esc(h)}</th>' for h in headers)
    rows_html = ''
    for i, row in enumerate(rows):
        cls = 'even' if i % 2 == 0 else 'odd'
        cells = ''.join(f'<td>{esc(v)}</td>' for v in row)
        rows_html += f'<tr class="{cls}">{cells}</tr>\n'

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<title>{esc(title)}</title>
<style>
  @page {{ size: A4 landscape; margin: 10mm; }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; font-size: 9pt; color: #000; }}
  h2 {{ text-align: center; font-size: 13pt; margin-bottom: 2mm; }}
  .meta {{ text-align: center; font-size: 8pt; color: #555; margin-bottom: 4mm; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 8.5pt; }}
  thead th {{ background: #2c3e50; color: #fff; padding: 2mm 2mm;
              text-align: left; border: 0.3pt solid #000; font-size: 8.5pt; }}
  tbody td {{ padding: 1.5mm 2mm; border: 0.3pt solid #ccc; vertical-align: middle; }}
  tr.even {{ background: #f7f7f7; }}
  tr.odd  {{ background: #ffffff; }}
  tfoot td {{ padding: 2mm; border-top: 1pt solid #000; font-weight: bold; }}
  .print-btn {{ display: block; margin: 5mm auto; padding: 2mm 10mm;
                font-size: 11pt; background: #2c3e50; color: white;
                border: none; border-radius: 3px; cursor: pointer; }}
  @media print {{ .print-btn {{ display: none; }} }}
</style>
</head>
<body>
<h2>{esc(title)}</h2>
<div class="meta">Generated: {date_str} &nbsp;|&nbsp; Total Records: {len(rows)}</div>
<table>
  <thead><tr>{header_html}</tr></thead>
  <tbody>{rows_html}</tbody>
</table>
<button class="print-btn" onclick="window.print()">&#128424; Print / Save as PDF</button>
</body>
</html>"""
